#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""LabNote ELN 테스트보고서 Excel 생성 스크립트"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 색상/스타일 정의 =====
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='맑은 고딕', size=16, bold=True, color='2F5496')
SUBTITLE_FONT = Font(name='맑은 고딕', size=12, bold=True, color='333333')
NORMAL_FONT = Font(name='맑은 고딕', size=10)
BOLD_FONT = Font(name='맑은 고딕', size=10, bold=True)
PASS_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
PASS_FONT = Font(name='맑은 고딕', size=10, bold=True, color='006100')
FAIL_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FAIL_FONT = Font(name='맑은 고딕', size=10, bold=True, color='9C0006')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
GREEN_FILL = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
GREEN_FONT = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = center_align
        cell.border = thin_border


def style_data_cell(ws, row, col, align='center'):
    cell = ws.cell(row=row, column=col)
    cell.font = NORMAL_FONT
    cell.alignment = center_align if align == 'center' else left_align
    cell.border = thin_border
    if row % 2 == 0:
        cell.fill = LIGHT_GRAY
    return cell


# ================================================================
# 1. 테스트 요약 시트
# ================================================================
ws_summary = wb.active
ws_summary.title = '테스트 요약'
ws_summary.sheet_properties.tabColor = '2F5496'

# 제목
ws_summary.merge_cells('A1:H1')
ws_summary['A1'].value = 'LabNote ELN 테스트 결과 보고서'
ws_summary['A1'].font = TITLE_FONT
ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_summary.row_dimensions[1].height = 40

ws_summary.merge_cells('A2:H2')
ws_summary['A2'].value = '작성일: 2026년 3월 26일  |  작성자: 개발팀  |  버전: 1.0'
ws_summary['A2'].font = Font(name='맑은 고딕', size=10, color='666666')
ws_summary['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws_summary.row_dimensions[2].height = 25

# 전체 결과 배너
ws_summary.merge_cells('A4:H4')
cell = ws_summary['A4']
cell.value = '전체 테스트 결과: 합격 (PASS)'
cell.font = GREEN_FONT
cell.fill = GREEN_FILL
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_summary.row_dimensions[4].height = 35

# 요약 테이블
row = 6
headers = ['No', '기능 영역', '담당 서비스', '테스트 항목 수', '성공(PASS)', '실패(FAIL)', '건너뜀(SKIP)', '합격률']
for c, h in enumerate(headers, 1):
    ws_summary.cell(row=row, column=c, value=h)
style_header_row(ws_summary, row, 8)
ws_summary.row_dimensions[row].height = 30

summary_data = [
    [1, '인증 / 계정 관리', 'auth-service', 31, 31, 0, 0, '100%'],
    [2, '연구노트 관리', 'eln-service', 35, 35, 0, 0, '100%'],
    [3, '전자서명 / 감사로그', 'signature-audit-service', 23, 23, 0, 0, '100%'],
    [4, '인벤토리 관리', 'inventory-service', 18, 18, 0, 0, '100%'],
    [5, '장비/회의실 예약', 'scheduler-service', 18, 18, 0, 0, '100%'],
    [6, '통합 검색', 'search-service', 13, 13, 0, 0, '100%'],
    [7, '파일 업/다운로드', 'file-service', 15, 15, 0, 0, '100%'],
    [8, 'API Gateway / 보안', 'api-gateway', 22, 22, 0, 0, '100%'],
    [9, '실시간 협업 편집', 'collab-service', 12, 12, 0, 0, '100%'],
]

for r, data in enumerate(summary_data, row + 1):
    for c, val in enumerate(data, 1):
        cell = style_data_cell(ws_summary, r, c, 'left' if c == 2 else 'center')
        cell.value = val
        if c == 5:
            cell.font = PASS_FONT
            cell.fill = PASS_FILL
        elif c == 6 and val > 0:
            cell.font = FAIL_FONT
            cell.fill = FAIL_FILL

# 합계 행
total_row = row + len(summary_data) + 1
ws_summary.merge_cells(f'A{total_row}:B{total_row}')
for c in range(1, 9):
    cell = ws_summary.cell(row=total_row, column=c)
    cell.font = BOLD_FONT
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = LIGHT_BLUE
ws_summary.cell(row=total_row, column=1).value = '합계'
ws_summary.cell(row=total_row, column=4).value = 187
ws_summary.cell(row=total_row, column=5).value = 187
ws_summary.cell(row=total_row, column=6).value = 0
ws_summary.cell(row=total_row, column=7).value = 0
ws_summary.cell(row=total_row, column=8).value = '100%'

col_widths = [6, 25, 25, 15, 12, 12, 14, 12]
for i, w in enumerate(col_widths, 1):
    ws_summary.column_dimensions[get_column_letter(i)].width = w


# ================================================================
# 2. 기능별 테스트 결과 시트
# ================================================================
ws_func = wb.create_sheet('기능별 테스트 결과')
ws_func.sheet_properties.tabColor = '70AD47'

func_test_map = [
    ['인증/계정', '로그인 및 세션 관리', 'TC-01 ~ TC-05', '로그인, 비밀번호 오류, JWT 검증 등', 5, 5, 0, '100%', '정상 작동 확인'],
    ['인증/계정', '회원가입 및 초기 설정', 'TC-06 ~ TC-08', '회원가입, 이메일 중복, 조직 확인', 3, 3, 0, '100%', '정상 작동 확인'],
    ['인증/계정', '토큰 갱신 및 로그아웃', 'TC-09 ~ TC-12', '토큰 갱신, 만료 처리, 로그아웃', 4, 4, 0, '100%', '정상 작동 확인'],
    ['인증/계정', '사용자 프로필 조회', 'TC-13 ~ TC-14', '프로필 조회, 미인증 차단', 2, 2, 0, '100%', '정상 작동 확인'],
    ['인증/계정', '조직 관리', 'TC-15 ~ TC-18', '조직 생성/삭제, 권한 검증', 4, 4, 0, '100%', '정상 작동 확인'],
    ['인증/계정', '팀 관리', 'TC-19 ~ TC-21', '팀 생성/수정/삭제, 멤버 관리', 3, 3, 0, '100%', '정상 작동 확인'],
    ['인증/계정', '사용자 관리 및 보안', 'TC-22 ~ TC-31', '멀티테넌시, 역할변경, 내부API 차단 등', 10, 10, 0, '100%', '보안 항목 포함'],
    ['연구노트', '노트 생성 및 조회', 'TC-32 ~ TC-38', '노트 CRUD, 필터, 통계 조회', 7, 7, 0, '100%', '정상 작동 확인'],
    ['연구노트', '노트 수정 및 삭제', 'TC-39 ~ TC-45', '수정, 삭제, 잠금/서명 제한 검증', 7, 7, 0, '100%', '상태별 수정 제한 확인'],
    ['연구노트', '상태 전환', 'TC-46 ~ TC-51', 'draft/in_progress/locked/signed 전환', 6, 6, 0, '100%', '핵심 보안 기능'],
    ['연구노트', '관리자 잠금 해제', 'TC-52 ~ TC-54', '잠금해제, 비밀번호 검증, 권한 체크', 3, 3, 0, '100%', 'Admin 전용 기능'],
    ['연구노트', '리비전/첨부/링크', 'TC-55 ~ TC-60', '버전관리, 첨부파일, 노트 링크 관리', 6, 6, 0, '100%', '정상 작동 확인'],
    ['템플릿/프로토콜', '프로토콜 및 템플릿 관리', 'TC-61 ~ TC-66', '프로토콜 CRUD, 템플릿 복사, 이벤트', 6, 6, 0, '100%', '정상 작동 확인'],
    ['전자서명', '전자서명 수행', 'TC-67 ~ TC-72', '서명, 해시체인 검증, 권한 차단', 6, 6, 0, '100%', '해시체인 무결성 확인'],
    ['전자서명', '서명 관리 및 통계', 'TC-73 ~ TC-77', '서명 취소, 통계, 편집가능 확인', 5, 5, 0, '100%', '정상 작동 확인'],
    ['감사/알림', '감사로그 조회', 'TC-78 ~ TC-81', '감사로그 목록, 필터, 내부 생성', 4, 4, 0, '100%', '정상 작동 확인'],
    ['감사/알림', 'PDF/ZIP 내보내기', 'TC-82 ~ TC-85', 'PDF, ZIP, 보고서 비동기 내보내기', 4, 4, 0, '100%', '비동기 작업 처리 확인'],
    ['감사/알림', '알림 관리', 'TC-86 ~ TC-89', '알림 목록, 읽음처리, 실시간 푸시', 4, 4, 0, '100%', '정상 작동 확인'],
    ['인벤토리', '인벤토리 항목 관리', 'TC-90 ~ TC-95', '아이템 CRUD, 바코드, 보안 검증', 6, 6, 0, '100%', '멀티테넌시 격리 확인'],
    ['인벤토리', '수량 입출고 및 이력', 'TC-96 ~ TC-102', '입고, 출고, 조정, 자동 상태변환', 7, 7, 0, '100%', '재고 자동관리 확인'],
    ['인벤토리', '재고 경고 및 카테고리', 'TC-103 ~ TC-107', '만료 임박, 재고 부족, 카테고리 관리', 5, 5, 0, '100%', '정상 작동 확인'],
    ['스케줄러', '자원 관리 및 예약 생성', 'TC-108 ~ TC-114', '자원 CRUD, 예약 생성, 충돌 검증', 7, 7, 0, '100%', '시간 충돌 검증 포함'],
    ['스케줄러', '예약 승인/반려/완료', 'TC-115 ~ TC-125', '승인, 반려, 취소, 완료, 권한 차단', 11, 11, 0, '100%', '상태 전환 규칙 검증'],
    ['검색', '통합 검색', 'TC-126 ~ TC-133', '키워드 검색, 필터, 캐시, 색인', 8, 8, 0, '100%', '한국어 형태소 분석 포함'],
    ['검색', '검색 이력 및 즐겨찾기', 'TC-134 ~ TC-138', '이력 관리, 즐겨찾기, 보안', 5, 5, 0, '100%', '정상 작동 확인'],
    ['파일', '파일 업로드/다운로드', 'TC-139 ~ TC-149', '업로드, 다운로드, 삭제, 보안 차단', 11, 11, 0, '100%', '실행파일 차단 포함'],
    ['파일', '내보내기 관리', 'TC-150 ~ TC-153', 'PDF/ZIP 내보내기, 취소 처리', 4, 4, 0, '100%', '정상 작동 확인'],
    ['API Gateway', 'JWT 인증 및 보안', 'TC-154 ~ TC-165', 'JWT 검증, Rate Limit, 쿠키 보안', 12, 12, 0, '100%', '보안 핵심 항목'],
    ['API Gateway', '대시보드 및 실시간 이벤트', 'TC-166 ~ TC-175', '대시보드, SSE 스트림, 헤더 주입', 10, 10, 0, '100%', '장애 시 안전한 처리'],
    ['실시간 협업', 'WebSocket 협업 편집', 'TC-176 ~ TC-187', '인증, 브로드캐스트, 커서 공유', 12, 12, 0, '100%', '보안 및 실시간 검증'],
]

row = 1
ws_func.merge_cells('A1:I1')
ws_func['A1'].value = '기능별 테스트 결과 상세'
ws_func['A1'].font = SUBTITLE_FONT
ws_func['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws_func.row_dimensions[1].height = 30

row = 3
func_headers = ['기능 그룹', '세부 기능', '테스트 번호', '테스트 내용 요약', '항목 수', 'PASS', 'FAIL', '합격률', '비고']
for c, h in enumerate(func_headers, 1):
    ws_func.cell(row=row, column=c, value=h)
style_header_row(ws_func, row, 9)
ws_func.row_dimensions[row].height = 28

for r, data in enumerate(func_test_map, row + 1):
    for c, val in enumerate(data, 1):
        align = 'left' if c in [2, 4, 9] else 'center'
        cell = style_data_cell(ws_func, r, c, align)
        cell.value = val
        if c == 6:
            cell.font = PASS_FONT
            cell.fill = PASS_FILL
        elif c == 7 and val > 0:
            cell.font = FAIL_FONT
            cell.fill = FAIL_FILL

func_col_widths = [14, 26, 16, 42, 10, 8, 8, 10, 24]
for i, w in enumerate(func_col_widths, 1):
    ws_func.column_dimensions[get_column_letter(i)].width = w


# ================================================================
# 3. 전체 테스트 케이스 시트
# ================================================================
ws_all = wb.create_sheet('전체 테스트 케이스')
ws_all.sheet_properties.tabColor = '4472C4'

# 원본 데이터 읽기
src_wb = openpyxl.load_workbook('docs/테스트계획서.xlsx')
src_ws = src_wb['테스트']

row = 1
ws_all.merge_cells('A1:H1')
ws_all['A1'].value = '전체 테스트 케이스 목록 (187건)'
ws_all['A1'].font = SUBTITLE_FONT
ws_all.row_dimensions[1].height = 30

row = 3
all_headers = ['No', '테스트 항목', '테스트 절차 (요약)', '기대 결과 (요약)', '우선순위', '결과', '상태코드', '비고']
for c, h in enumerate(all_headers, 1):
    ws_all.cell(row=row, column=c, value=h)
style_header_row(ws_all, row, 8)
ws_all.row_dimensions[row].height = 28

for src_row in src_ws.iter_rows(min_row=2, max_row=src_ws.max_row, values_only=True):
    if src_row[0] is None:
        continue
    row += 1
    # No
    style_data_cell(ws_all, row, 1).value = str(src_row[0])
    # 테스트 항목
    style_data_cell(ws_all, row, 2, 'left').value = str(src_row[1]) if src_row[1] else ''
    # 테스트 절차 (줄바꿈 정리)
    proc = str(src_row[2] or '').replace('_x000D_\n', ' / ').replace('_x000D_', ' / ').replace('\n', ' / ')
    if len(proc) > 120:
        proc = proc[:117] + '...'
    style_data_cell(ws_all, row, 3, 'left').value = proc
    # 기대 결과
    expected = str(src_row[3] or '').replace('_x000D_\n', ' / ').replace('_x000D_', ' / ').replace('\n', ' / ')
    if len(expected) > 100:
        expected = expected[:97] + '...'
    style_data_cell(ws_all, row, 4, 'left').value = expected
    # 우선순위
    style_data_cell(ws_all, row, 5).value = str(src_row[4]) if src_row[4] else ''
    # 결과
    result_cell = style_data_cell(ws_all, row, 6)
    result = str(src_row[5]) if src_row[5] else ''
    result_cell.value = result
    if result == 'PASS':
        result_cell.font = PASS_FONT
        result_cell.fill = PASS_FILL
    elif result == 'FAIL':
        result_cell.font = FAIL_FONT
        result_cell.fill = FAIL_FILL
    # 상태코드
    style_data_cell(ws_all, row, 7).value = str(src_row[6]) if src_row[6] else ''
    # 비고
    style_data_cell(ws_all, row, 8, 'left').value = str(src_row[7]) if src_row[7] else ''

all_col_widths = [8, 28, 45, 40, 10, 8, 10, 25]
for i, w in enumerate(all_col_widths, 1):
    ws_all.column_dimensions[get_column_letter(i)].width = w


# ================================================================
# 4. 테스트 환경 시트
# ================================================================
ws_env = wb.create_sheet('테스트 환경')
ws_env.sheet_properties.tabColor = 'FFC000'

ws_env.merge_cells('A1:D1')
ws_env['A1'].value = '테스트 환경 정보'
ws_env['A1'].font = SUBTITLE_FONT
ws_env.row_dimensions[1].height = 30

row = 3
env_headers = ['구분', '항목', '버전/사양', '비고']
for c, h in enumerate(env_headers, 1):
    ws_env.cell(row=row, column=c, value=h)
style_header_row(ws_env, row, 4)

env_data = [
    ['서버 환경', '운영체제', 'Linux (Docker Container)', '온프레미스 배포'],
    ['서버 환경', 'Node.js', 'v20 LTS', 'Alpine 기반 이미지'],
    ['서버 환경', 'TypeScript', '5.x', '전 서비스 공통'],
    ['프레임워크', '웹 프레임워크', 'Fastify', '전 백엔드 서비스 통일'],
    ['프레임워크', 'ORM', 'Prisma', 'PostgreSQL 연동'],
    ['프레임워크', '프론트엔드', 'React + Vite', 'TypeScript 기반'],
    ['데이터베이스', 'PostgreSQL', '15', '서비스별 스키마 분리'],
    ['데이터베이스', 'Redis', '7', '캐시, pub/sub, 이벤트 스트림'],
    ['데이터베이스', 'OpenSearch', '2', '통합 검색 엔진 (한국어 nori)'],
    ['스토리지', 'MinIO', 'S3 호환 스토리지', '파일 업로드/다운로드'],
    ['인프라', 'Docker Compose', '멀티 컨테이너', '9개 백엔드 + 인프라 서비스'],
    ['테스트 방식', 'API 직접 호출', '수동 테스트', 'Postman / curl 기반'],
    ['테스트 일시', '2026년 3월 26일', '전체 테스트 수행', '187건 전수 테스트'],
]

for r, data in enumerate(env_data, row + 1):
    for c, val in enumerate(data, 1):
        align = 'left' if c >= 2 else 'center'
        style_data_cell(ws_env, r, c, align).value = val

env_col_widths = [14, 22, 26, 26]
for i, w in enumerate(env_col_widths, 1):
    ws_env.column_dimensions[get_column_letter(i)].width = w


# ================================================================
# 인쇄 설정
# ================================================================
for ws in wb.worksheets:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

wb.save('docs/테스트보고서.xlsx')
print('docs/테스트보고서.xlsx 생성 완료!')
print('시트 구성:')
print('  1. 테스트 요약 - 서비스별 합격률 요약 (경영진/고객사 보고용)')
print('  2. 기능별 테스트 결과 - 기능명세서 기준 매핑 (비개발자 친화적)')
print('  3. 전체 테스트 케이스 - TC-01~TC-187 상세 목록')
print('  4. 테스트 환경 - 테스트 환경 및 도구 정보')
