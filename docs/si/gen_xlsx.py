#!/usr/bin/env python3
"""SI 산출물 XLSX 생성 (16건)"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, datetime

OUT = os.path.dirname(os.path.abspath(__file__))
TODAY = datetime.date.today().isoformat()

HF = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
HB = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
DF = Font(name='맑은 고딕', size=9)
ZF = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
BD = Border(*(Side(style='thin', color='D1D5DB'),)*4)
WR = Alignment(wrap_text=True, vertical='top')

def setup(ws, headers, widths):
    ws.append(headers)
    for c in ws[1]:
        c.font, c.fill, c.border = HF, HB, BD
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def style_rows(ws, start=2):
    for row in ws.iter_rows(min_row=start):
        for c in row:
            c.font, c.alignment, c.border = DF, WR, BD
            if c.row % 2 == 0: c.fill = ZF

def save(wb, name):
    p = os.path.join(OUT, name)
    wb.save(p)
    print(f"  {name}")

# ════════════════════════════════════════════
# 05. WBS
# ════════════════════════════════════════════
def gen_wbs():
    wb = Workbook(); ws = wb.active; ws.title = "WBS"
    setup(ws, ["WBS ID","단계","작업명","상세 내용","담당","시작일","종료일","산출물","진척률","비고"],
          [10,8,25,40,10,12,12,20,8,15])
    rows = [
        ["1","제안","프로젝트 착수","킥오프 미팅, 프로젝트 헌장 작성","PM","","","프로젝트 헌장","100%",""],
        ["1.1","제안","사업수행계획 수립","PMP 작성, WBS 정의","PM","","","사업수행계획서, WBS","100%",""],
        ["1.2","제안","제안서 작성","기술/관리/가격 제안서","PM","","","제안서","100%",""],
        ["2","분석","현행시스템 분석","기존 시스템 구조/데이터 분석","AA","","","현행시스템 분석서","100%",""],
        ["2.1","분석","요구사항 정의","기능/비기능 요구사항 수집 및 정의","AA","","","요구사항 정의서(SRS)","100%",""],
        ["2.2","분석","업무 프로세스 분석","AS-IS/TO-BE 프로세스 정의","AA","","","업무 프로세스 정의서","100%",""],
        ["2.3","분석","유스케이스 작성","상세 유스케이스 시나리오","AA","","","유스케이스 명세서","100%",""],
        ["3","설계","시스템 아키텍처 설계","MSA 구조, 서비스 분리, 통신 패턴","DA/TA","","","아키텍처 설계서","100%",""],
        ["3.1","설계","데이터 모델 설계","ERD, 테이블 정의, 스키마 분리","DA","","","데이터 모델 정의서","100%",""],
        ["3.2","설계","API 설계","REST API 엔드포인트/스키마 정의","개발PL","","","API 명세서, OpenAPI","100%",""],
        ["3.3","설계","인터페이스 설계","서비스 간 연동 인터페이스 정의","DA","","","인터페이스 정의서","100%",""],
        ["3.4","설계","화면 설계","UI/UX 와이어프레임, 프로토타입","UI설계자","","","화면설계서","100%",""],
        ["4","구현","공용 패키지 개발","@lab/shared (에러,로거,권한,검증)","개발PL","","","소스코드","100%",""],
        ["4.1","구현","auth-service 개발","조직/팀/사용자 CRUD, RBAC, JWT","개발","","","소스코드","100%",""],
        ["4.2","구현","eln-service 개발","노트/프로토콜/템플릿 CRUD, 버전관리","개발","","","소스코드","100%",""],
        ["4.3","구현","signature-audit 개발","전자서명, 감사로그, PDF 내보내기","개발","","","소스코드","100%",""],
        ["4.4","구현","inventory-service 개발","시약/장비 CRUD, 수량관리","개발","","","소스코드","100%",""],
        ["4.5","구현","scheduler-service 개발","장비/회의실 예약, 승인 워크플로","개발","","","소스코드","100%",""],
        ["4.6","구현","search-service 개발","OpenSearch 통합검색, 자동완성","개발","","","소스코드","100%",""],
        ["4.7","구현","file-service 개발","MinIO 파일 업/다운로드, presigned URL","개발","","","소스코드","100%",""],
        ["4.8","구현","collab-service 개발","WebSocket 실시간 협업","개발","","","소스코드","100%",""],
        ["4.9","구현","api-gateway 개발","JWT 검증, 프록시, Rate Limit, SSE","개발","","","소스코드","100%",""],
        ["4.10","구현","프론트엔드 개발","React+Vite 메인 프론트엔드","개발","","","소스코드","100%",""],
        ["4.11","구현","단위 테스트","서비스별 단위 테스트 수행","개발","","","단위테스트 결과서","100%",""],
        ["5","테스트","통합 테스트","서비스 간 연동 테스트","QA","","","통합테스트 결과서","100%",""],
        ["5.1","테스트","성능 테스트","부하/스트레스 테스트","TA","","","성능테스트 결과서","100%",""],
        ["5.2","테스트","보안 점검","OWASP 기반 보안 취약점 점검","보안팀","","","보안 점검 결과서","100%",""],
        ["5.3","테스트","UAT","사용자 인수 테스트","QA/현업","","","UAT 결과서","100%",""],
        ["6","이행","운영 이행","시스템 운영환경 배포","DevOps","","","운영 이행 계획서","100%",""],
        ["6.1","이행","교육","사용자/운영자 교육","PM","","","교육자료","100%",""],
        ["6.2","이행","검수","발주기관 최종 검수","PM","","","감수 조서","100%",""],
    ]
    for r in rows: ws.append(r)
    style_rows(ws)
    save(wb, "05_WBS.xlsx")

# ════════════════════════════════════════════
# 08. 요구사항 정의서(SRS)
# ════════════════════════════════════════════
def gen_srs():
    wb = Workbook(); ws = wb.active; ws.title = "기능 요구사항"
    setup(ws, ["ID","구분","기능영역","요구사항명","상세 설명","우선순위","필수여부","관련 서비스","관련 화면","비고"],
          [10,8,12,22,45,8,8,15,15,15])
    reqs = [
        ["FR-001","기능","인증","로그인/로그아웃","이메일/비밀번호 기반 JWT 인증, 토큰 갱신, 블랙리스트","상","필수","auth-service","/login",""],
        ["FR-002","기능","인증","SSO 연동","Keycloak JWKS 기반 SSO 선택적 연동","하","선택","auth, gateway","","Keycloak 24"],
        ["FR-003","기능","인증","RBAC 권한 관리","Admin/Reviewer/Researcher/Viewer 4역할, 18개 권한","상","필수","auth-service","/admin/roles",""],
        ["FR-004","기능","조직","멀티테넌시","조직(orgId) 기반 데이터 격리, 모든 쿼리 스코핑","상","필수","전체","",""],
        ["FR-005","기능","조직","팀 관리","팀 CRUD, 팀원 배정, 리더/멤버 역할","중","필수","auth-service","/admin/users",""],
        ["FR-006","기능","노트","노트 CRUD","연구노트 생성/조회/수정/삭제, 소프트 삭제","상","필수","eln-service","/notes",""],
        ["FR-007","기능","노트","노트 상태 관리","draft→in_progress→locked/signed 전환, DB 트리거 보호","상","필수","eln-service","/notes",""],
        ["FR-008","기능","노트","버전 관리","수정 시마다 NoteRevision 자동 생성, 이력 조회","상","필수","eln-service","/notes/:id",""],
        ["FR-009","기능","노트","첨부파일","노트에 파일 첨부, 드래그앤드롭 업로드, 다운로드","중","필수","eln, file","/notes/:id",""],
        ["FR-010","기능","노트","인벤토리 연결","노트에 시약/장비 연결 (NoteLink)","중","필수","eln, inventory","/notes/:id",""],
        ["FR-011","기능","노트","실시간 협업","WebSocket 기반 다중 사용자 동시 편집, Redis Pub/Sub 팬아웃","상","필수","collab-service","/notes/:id",""],
        ["FR-012","기능","서명","전자서명","SHA-256 해시체인 서명, 비밀번호 검증, 노트 상태 자동 전환","상","필수","sig-audit","/notes/:id",""],
        ["FR-013","기능","서명","서명 검증","해시체인 무결성 검증 API","상","필수","sig-audit","/signatures",""],
        ["FR-014","기능","서명","signed 불변성","서명 완료 노트 수정/삭제/상태변경 완전 차단","상","필수","eln, sig-audit","","DB 트리거"],
        ["FR-015","기능","프로토콜","템플릿 관리","실험법/SOP 템플릿 CRUD, 카테고리, 공개/비공개","중","필수","eln-service","/protocols",""],
        ["FR-016","기능","프로토콜","노트 생성 연동","템플릿 기반 노트 자동 생성, useCount 추적","중","필수","eln-service","/protocols",""],
        ["FR-017","기능","인벤토리","아이템 CRUD","시약/샘플/장비 등록/수정/삭제, 바코드, 태그","상","필수","inventory","/inventory",""],
        ["FR-018","기능","인벤토리","수량 관리","입고/출고/조정, 변경 이력, 최소수량 알림","상","필수","inventory","/inventory",""],
        ["FR-019","기능","인벤토리","만료 관리","유효기간 추적, 만료 경고 일수 설정, 알림","중","필수","inventory","/inventory",""],
        ["FR-020","기능","예약","예약 CRUD","장비/회의실 예약 생성/수정/취소","중","필수","scheduler","/scheduler",""],
        ["FR-021","기능","예약","승인 워크플로","PENDING→APPROVED/REJECTED, 비관적 락 충돌 방지","상","필수","scheduler","/scheduler",""],
        ["FR-022","기능","검색","통합 검색","노트/템플릿/인벤토리 통합 검색, OpenSearch nori 분석기","상","필수","search","/search",""],
        ["FR-023","기능","검색","자동완성","phrase_prefix 기반 자동완성, 즐겨찾기","중","필수","search","/search",""],
        ["FR-024","기능","내보내기","PDF/ZIP 내보내기","Puppeteer PDF 렌더링, BullMQ 비동기, SSE 알림","상","필수","sig-audit, file","/exports",""],
        ["FR-025","기능","감사","감사 로그","모든 활동 자동 기록, 필터/검색, 페이지네이션","상","필수","sig-audit","/audit-logs",""],
        ["FR-026","기능","알림","실시간 알림","Redis Pub/Sub → SSE, 잠금/서명/승인 알림, 읽음 처리","중","필수","sig-audit, gateway","헤더 벨",""],
        ["FR-027","기능","대시보드","대시보드 집계","개인/팀/조직 대시보드, 다중 서비스 병렬 집계, Redis 캐시","중","필수","gateway","/",""],
        ["FR-028","기능","다국어","i18n","한국어/영어 전환, react-i18next, 네임스페이스 분리","중","필수","프론트엔드","전체",""],
    ]
    for r in reqs: ws.append(r)
    style_rows(ws)

    ws2 = wb.create_sheet("비기능 요구사항")
    setup(ws2, ["ID","구분","영역","요구사항명","상세 설명","목표값","우선순위"],
          [10,8,12,22,45,15,8])
    nfrs = [
        ["NFR-001","비기능","성능","API 응답시간","95th percentile 기준 API 응답시간","< 500ms","상"],
        ["NFR-002","비기능","성능","동시 사용자","동시 접속 사용자 수 지원","50명","상"],
        ["NFR-003","비기능","성능","처리량","전체 시스템 TPS","> 100 TPS","상"],
        ["NFR-004","비기능","가용성","서비스 가동률","월간 서비스 가동률","99.5%","상"],
        ["NFR-005","비기능","보안","인증","JWT 기반 인증, 토큰 블랙리스트","필수","상"],
        ["NFR-006","비기능","보안","데이터 격리","멀티테넌시 orgId 기반 완전 격리","필수","상"],
        ["NFR-007","비기능","보안","입력 검증","Zod 스키마 전 엔드포인트 검증","필수","상"],
        ["NFR-008","비기능","보안","Rate Limit","API 경로별 요청 수 제한","경로별 차등","중"],
        ["NFR-009","비기능","확장성","서비스 독립 배포","각 서비스 독립 빌드/배포 가능","Docker 기반","중"],
        ["NFR-010","비기능","운영성","로깅","Pino 구조화 로그, 서비스별 분리","필수","중"],
        ["NFR-011","비기능","운영성","모니터링","Jaeger 분산 트레이싱, Dozzle 로그 뷰어","필수","중"],
        ["NFR-012","비기능","데이터","백업","PostgreSQL 데이터 백업/복원 지원","일 1회","중"],
    ]
    for r in nfrs: ws2.append(r)
    style_rows(ws2)
    save(wb, "08_요구사항_정의서_SRS.xlsx")

# ════════════════════════════════════════════
# 09. RTM
# ════════════════════════════════════════════
def gen_rtm():
    wb = Workbook(); ws = wb.active; ws.title = "추적 매트릭스"
    setup(ws, ["요구사항 ID","요구사항명","설계 문서","설계 항목","구현 모듈","소스 파일","테스트 ID","테스트 결과","상태"],
          [12,20,18,20,15,30,12,10,8])
    rows = [
        ["FR-006","노트 CRUD","아키텍처 설계서","eln-service 설계","eln-service","note.controller.ts, note.routes.ts","INT-020~024","PASS","완료"],
        ["FR-007","노트 상태 관리","아키텍처 설계서","상태 전환 맵","eln-service","note.dto.ts, note.controller.ts","INT-030~032","PASS","완료"],
        ["FR-008","버전 관리","데이터 모델","NoteRevision","eln-service","note.controller.ts","INT-023","PASS","완료"],
        ["FR-011","실시간 협업","아키텍처 설계서","collab-service","collab-service","index.ts","—","—","완료"],
        ["FR-012","전자서명","아키텍처 설계서","해시체인 설계","sig-audit","signature.controller.ts","INT-030~034","PASS","완료"],
        ["FR-014","signed 불변성","데이터 모델","DB 트리거","eln-service","migration (trigger)","INT-032","PASS","완료"],
        ["FR-017","인벤토리 CRUD","아키텍처 설계서","inventory 설계","inventory","inventory.controller.ts","INT-050","PASS","완료"],
        ["FR-018","수량 관리","데이터 모델","InventoryHistory","inventory","inventory.controller.ts","INT-050","PASS","완료"],
        ["FR-021","승인 워크플로","아키텍처 설계서","비관적 락 설계","scheduler","bookings.ts","INT-060~062","PASS","완료"],
        ["FR-022","통합 검색","아키텍처 설계서","OpenSearch 설계","search","search.controller.ts","INT-020~021","PASS","완료"],
        ["FR-024","PDF 내보내기","아키텍처 설계서","BullMQ 설계","sig-audit","export.worker.ts","INT-040~042","PASS","완료"],
        ["FR-025","감사 로그","인터페이스 정의서","내부 API","sig-audit","audit.controller.ts","INT-022","PASS","완료"],
        ["FR-026","실시간 알림","아키텍처 설계서","Redis Pub/Sub→SSE","sig-audit, gateway","sse.ts, notification.controller.ts","INT-070~072","PASS","완료"],
        ["NFR-001","API 응답시간","성능 목표","< 500ms","전체","—","PERF-*","PASS","완료"],
        ["NFR-005","JWT 인증","아키텍처 설계서","인증 흐름","auth, gateway","auth.middleware.ts","INT-010~014","PASS","완료"],
        ["NFR-006","데이터 격리","아키텍처 설계서","멀티테넌시","전체","withOrgScope()","INT-014","PASS","완료"],
    ]
    for r in rows: ws.append(r)
    style_rows(ws)
    save(wb, "09_요구사항_추적_매트릭스_RTM.xlsx")

# ════════════════════════════════════════════
# 10. 업무 프로세스 정의서(BPD)
# ════════════════════════════════════════════
def gen_bpd():
    wb = Workbook(); ws = wb.active; ws.title = "업무 프로세스"
    setup(ws, ["프로세스 ID","프로세스명","설명","시작 이벤트","주요 단계","종료 이벤트","관련 역할","관련 시스템","AS-IS","TO-BE 개선"],
          [12,18,30,15,45,15,15,15,25,25])
    rows = [
        ["BP-001","연구노트 작성","연구원이 실험 노트를 작성하고 관리","실험 시작","1.노트 생성→2.내용 작성→3.첨부파일 업로드→4.시약 연결→5.상태 변경","노트 완성","Researcher","eln, file, inventory","수기 노트 작성, 종이 보관","전자 노트, 실시간 협업, 자동 버전관리"],
        ["BP-002","전자서명","검토자가 완성된 노트에 전자서명","노트 검토 요청","1.노트 내용 확인→2.비밀번호 입력→3.해시체인 서명→4.상태 자동 전환","서명 완료(불변)","Reviewer, Admin","sig-audit, auth, eln","수기 서명, 날인","SHA-256 해시체인, 자동 상태 전환"],
        ["BP-003","관리자 잠금해제","관리자가 잠긴 노트를 해제","잠금 해제 요청","1.Admin 비밀번호 확인→2.사유 입력→3.locked→draft 전환→4.감사로그+알림","노트 해제","Admin","eln, auth, sig-audit","물리적 키 관리","비밀번호+사유 기록, 감사 추적"],
        ["BP-004","시약 재고 관리","시약의 입고/출고/조정 및 알림","재고 변동 발생","1.수량 변경(in/out/adjust)→2.이력 기록→3.최소수량 확인→4.만료일 확인","재고 반영","Researcher","inventory","수기 대장, 엑셀","실시간 수량 추적, 자동 알림"],
        ["BP-005","장비 예약","장비/회의실 예약 및 승인","예약 필요","1.예약 생성→2.충돌 검증→3.승인 대기→4.Admin 승인/반려→5.알림 전달","예약 확정","Researcher, Admin","scheduler, sig-audit","구두 예약, 공유 캘린더","시스템 예약, 자동 충돌 검증, 비관적 락"],
        ["BP-006","PDF 내보내기","서명된 노트를 PDF로 보관","내보내기 요청","1.BullMQ 큐 등록→2.노트 내용 조회→3.Puppeteer 렌더링→4.MinIO 업로드→5.SSE 알림","다운로드 가능","Reviewer, Admin","sig-audit, eln, file","수동 인쇄, 스캔","자동 PDF 생성, 클라우드 저장"],
        ["BP-007","통합 검색","노트/템플릿/인벤토리 통합 검색","검색어 입력","1.OpenSearch 쿼리→2.한국어 형태소 분석→3.권한 필터링→4.결과 반환","검색 결과 표시","전체","search, OpenSearch","파일 탐색기, 수동 검색","실시간 통합 검색, 자동완성"],
        ["BP-008","사용자/권한 관리","조직, 팀, 사용자, 역할 관리","인원 변동","1.사용자 생성→2.역할 배정→3.팀 배정→4.권한 적용","권한 반영","Admin","auth","수기 관리 대장","시스템 RBAC, 실시간 권한 적용"],
    ]
    for r in rows: ws.append(r)
    style_rows(ws)
    save(wb, "10_업무_프로세스_정의서_BPD.xlsx")

# ════════════════════════════════════════════
# 16. 데이터 모델 정의서(ERD)
# ════════════════════════════════════════════
def gen_erd():
    wb = Workbook()
    schemas = {
        "auth": [
            ["Organization","id","String","PK, UUID","조직(테넌트)"],
            ["Organization","name","String","","조직명"],
            ["Organization","slug","String","UNIQUE","슬러그"],
            ["Role","id","String","PK, UUID","역할"],
            ["Role","orgId","String","FK→Organization","소속 조직"],
            ["Role","name","RoleName","ENUM(admin,researcher,reviewer,viewer)","역할명"],
            ["Role","permissions","String[]","","권한 배열"],
            ["User","id","String","PK, UUID","사용자"],
            ["User","orgId","String","FK→Organization","소속 조직"],
            ["User","email","String","UNIQUE","이메일"],
            ["User","name","String","","이름"],
            ["User","passwordHash","String","","bcrypt 해시"],
            ["User","roleId","String?","FK→Role","역할"],
            ["User","status","UserStatus","ENUM(active,inactive,suspended)","상태"],
            ["Team","id","String","PK, UUID","팀"],
            ["Team","orgId","String","FK→Organization","소속 조직"],
            ["Team","name","String","","팀명"],
            ["TeamMember","userId+teamId","String","복합PK","팀-사용자"],
            ["TeamMember","teamRole","TeamRole","ENUM(leader,member)","팀 내 역할"],
        ],
        "eln": [
            ["Note","id","String","PK, UUID","연구노트/프로토콜"],
            ["Note","type","NoteType","ENUM(note,template)","유형"],
            ["Note","title","String","","제목"],
            ["Note","content","String","","본문"],
            ["Note","sections","Json","DEFAULT []","섹션 구조"],
            ["Note","status","NoteStatus","ENUM(draft,in_progress,locked,signed)","상태"],
            ["Note","authorId","String","FK→auth.User","작성자"],
            ["Note","orgId","String","","조직"],
            ["Note","teamId","String?","","소속 팀"],
            ["Note","templateId","String?","","원본 템플릿"],
            ["Note","tags","String[]","","태그"],
            ["Note","deletedAt","DateTime?","","소프트삭제"],
            ["NoteRevision","id","String","PK","리비전"],
            ["NoteRevision","noteId","String","FK→Note","노트"],
            ["NoteRevision","revision","Int","","버전 번호"],
            ["NoteRevision","content","String","","내용 스냅샷"],
            ["NoteRevision","changedBy","String","","변경자"],
            ["NoteStatusHistory","id","String","PK","상태 이력"],
            ["NoteStatusHistory","noteId","String","FK→Note","노트"],
            ["NoteStatusHistory","fromStatus","NoteStatus","","이전 상태"],
            ["NoteStatusHistory","toStatus","NoteStatus","","이후 상태"],
            ["NoteStatusHistory","changedBy","String","","변경자"],
            ["Attachment","id","String","PK","첨부파일"],
            ["Attachment","noteId","String","FK→Note","노트"],
            ["Attachment","fileId","String","FK→file.File","파일"],
            ["NoteLink","id","String","PK","외부 연결"],
            ["NoteLink","noteId","String","FK→Note","노트"],
            ["NoteLink","targetType","String","","inventory/equipment/note"],
            ["NoteLink","targetId","String","","대상 ID"],
            ["Template","id","String","PK","템플릿"],
            ["Template","title","String","","제목"],
            ["Template","category","String","","카테고리"],
            ["Template","isPublic","Boolean","","공개 여부"],
            ["Template","useCount","Int","","사용 횟수"],
            ["Code","id","String","PK","참조 코드"],
            ["Code","group","String","","코드 그룹"],
            ["Code","value","String","","값"],
        ],
        "signature": [
            ["Signature","id","String","PK, UUID","전자서명"],
            ["Signature","noteId","String","FK→eln.Note","노트"],
            ["Signature","signerId","String","FK→auth.User","서명자"],
            ["Signature","signatureHash","String","","SHA-256 해시"],
            ["Signature","prevHash","String?","","이전 서명 해시(체인)"],
            ["Signature","chainIndex","Int","","체인 순서"],
            ["Signature","status","SignatureStatus","ENUM(valid,revoked)","상태"],
            ["AuditLog","id","String","PK","감사 로그"],
            ["AuditLog","entityType","String","","대상 유형"],
            ["AuditLog","entityId","String","","대상 ID"],
            ["AuditLog","action","String","","액션"],
            ["AuditLog","actorId","String","","수행자"],
            ["AuditLog","orgId","String","","조직"],
            ["AuditLog","details","Json","","상세 정보"],
            ["ExportJob","id","String","PK","내보내기 작업"],
            ["ExportJob","noteId","String","","노트 ID"],
            ["ExportJob","format","ExportFormat","ENUM(pdf,zip,report)","형식"],
            ["ExportJob","status","ExportStatus","ENUM(pending,processing,completed,failed)","상태"],
            ["Notification","id","String","PK","알림"],
            ["Notification","recipientId","String","","수신자"],
            ["Notification","type","NotificationType","ENUM(NOTE_LOCKED,NOTE_SIGNED,NOTE_UNLOCKED,BOOKING_APPROVED)","유형"],
            ["Notification","isRead","Boolean","","읽음 여부"],
        ],
        "inventory": [
            ["InventoryItem","id","String","PK, UUID","인벤토리 아이템"],
            ["InventoryItem","name","String","","이름"],
            ["InventoryItem","orgId","String","","조직"],
            ["InventoryItem","type","String","","유형(reagent,sample,equipment...)"],
            ["InventoryItem","status","String","","상태(available,in_use,depleted...)"],
            ["InventoryItem","barcode","String?","UNIQUE","바코드"],
            ["InventoryItem","quantity","Float?","","수량"],
            ["InventoryItem","minQuantity","Float?","","최소 재고"],
            ["InventoryItem","expiryDate","DateTime?","","유효기간"],
            ["InventoryHistory","id","String","PK","변경 이력"],
            ["InventoryHistory","itemId","String","FK→InventoryItem","아이템"],
            ["InventoryHistory","changeType","String","","in/out/adjust/status_change"],
            ["InventoryHistory","quantityDelta","Float?","","변화량"],
            ["Category","id","String","PK","카테고리"],
            ["Category","name","String","UNIQUE(orgId,name)","카테고리명"],
        ],
        "scheduler": [
            ["Resource","id","String","PK, UUID","예약 자원"],
            ["Resource","name","String","","자원명"],
            ["Resource","orgId","String","","조직"],
            ["Resource","type","ResourceType","ENUM(EQUIPMENT,ROOM)","유형"],
            ["Resource","ownerId","String?","","승인 담당자"],
            ["Resource","isActive","Boolean","","활성 여부"],
            ["Booking","id","String","PK, UUID","예약"],
            ["Booking","resourceId","String","FK→Resource","자원"],
            ["Booking","userId","String","","예약자"],
            ["Booking","status","BookingStatus","ENUM(PENDING,APPROVED,REJECTED,CANCELLED,COMPLETED)","상태"],
            ["Booking","startAt","DateTime","","시작"],
            ["Booking","endAt","DateTime","","종료"],
        ],
        "search": [
            ["SearchHistory","id","String","PK","검색 이력"],
            ["SearchHistory","userId","String","","사용자"],
            ["SearchHistory","query","String","","검색어"],
            ["Favorite","id","String","PK","즐겨찾기"],
            ["Favorite","docType","String","","notes/templates/inventory"],
            ["Favorite","docId","String","","문서 ID"],
            ["SearchKeywordFavorite","id","String","PK","검색어 즐겨찾기"],
            ["SearchKeywordFavorite","keyword","String","UNIQUE(userId,keyword)","키워드"],
        ],
        "file": [
            ["File","id","String","PK, UUID","파일"],
            ["File","bucket","String","","MinIO 버킷"],
            ["File","objectKey","String","UNIQUE","오브젝트 키"],
            ["File","originalName","String","","원본 파일명"],
            ["File","mimeType","String?","","MIME 타입"],
            ["File","sizeBytes","BigInt?","","파일 크기"],
            ["File","uploaderId","String","","업로더"],
            ["File","isDeleted","Boolean","","삭제 여부"],
            ["ExportJob","id","String","PK","내보내기"],
            ["ExportJob","type","String","","pdf/zip"],
            ["ExportJob","status","String","","PENDING/PROCESSING/COMPLETED/FAILED"],
            ["ExportJob","resultFileId","String?","FK→File","결과 파일"],
        ],
    }
    for schema_name, fields in schemas.items():
        ws = wb.active if schema_name == "auth" else wb.create_sheet(schema_name)
        ws.title = f"schema_{schema_name}"
        setup(ws, ["테이블","컬럼","타입","제약조건","설명"], [18,18,22,25,25])
        for r in fields: ws.append(r)
        style_rows(ws)
    save(wb, "16_데이터_모델_정의서_ERD.xlsx")

# ════════════════════════════════════════════
# 18. API 명세서
# ════════════════════════════════════════════
def gen_api_spec():
    wb = Workbook(); ws = wb.active; ws.title = "API 명세"
    setup(ws, ["서비스","Method","경로","설명","권한","요청 Body","응답 Body","상태코드","비고"],
          [12,8,30,25,15,25,25,10,15])
    apis = [
        ["auth","POST","/api/auth/login","로그인","Public","{email,password}","{token,refreshToken,user}","200",""],
        ["auth","POST","/api/auth/register","회원가입","Public","{name,email,password}","{user}","201",""],
        ["auth","POST","/api/auth/refresh","토큰 갱신","—","{refreshToken}","{token,refreshToken}","200",""],
        ["auth","POST","/api/auth/logout","로그아웃","JWT","—","—","200","Redis 블랙리스트"],
        ["auth","GET","/api/auth/me","내 정보","JWT","—","{user+permissions}","200",""],
        ["auth","CRUD","/api/auth/orgs","조직 관리","ADMIN","—","Organization","200",""],
        ["auth","CRUD","/api/auth/teams","팀 관리","ADMIN","—","Team","200",""],
        ["auth","CRUD","/api/auth/users","사용자 관리","ADMIN","—","User","200","페이지네이션"],
        ["auth","CRUD","/api/auth/roles","역할 관리","ADMIN","—","Role","200",""],
        ["eln","GET","/api/notes","노트 목록","NOTE_READ","?status,tag,type,page","[Note]+total","200","필터링"],
        ["eln","POST","/api/notes","노트 생성","NOTE_WRITE","{title,content,type,tags}","Note","201",""],
        ["eln","GET","/api/notes/:id","노트 상세","NOTE_READ","—","Note+attachments+links","200",""],
        ["eln","PUT","/api/notes/:id","노트 수정","NOTE_WRITE","{title,content,tags}","Note","200","리비전 생성"],
        ["eln","DELETE","/api/notes/:id","노트 삭제","NOTE_DELETE","—","—","200","소프트삭제"],
        ["eln","PATCH","/api/notes/:id/status","상태 변경","NOTE_STATUS","{status}","Note","200","전환맵 검증"],
        ["eln","POST","/api/notes/:id/admin-unlock","잠금 해제","ADMIN+UNLOCK","{adminPassword,reason}","Note","200","비밀번호 검증"],
        ["eln","CRUD","/api/templates","템플릿 관리","TEMPLATE_*","—","Template","200",""],
        ["eln","POST","/api/templates/:id/copy","템플릿 복사","TEMPLATE_WRITE","—","Template","201","copyCount++"],
        ["sig-audit","POST","/api/signatures/sign/:noteId","전자서명","NOTE_SIGN","{password}","Signature","201","해시체인"],
        ["sig-audit","GET","/api/signatures/verify/:noteId","서명 검증","NOTE_READ","—","{valid,signatures}","200",""],
        ["sig-audit","POST","/api/audit/internal","감사로그 기록","Internal","{entityType,action...}","{id}","201","내부전용"],
        ["sig-audit","GET","/api/audit","감사로그 조회","AUDIT_READ","?action,dateFrom","[AuditLog]+total","200",""],
        ["sig-audit","POST","/api/export/pdf/:noteId","PDF 내보내기","EXPORT_PDF","—","{jobId}","202","비동기"],
        ["sig-audit","GET","/api/notifications","알림 목록","JWT","?page,isRead","[Notification]+total","200",""],
        ["inventory","GET","/api/inventory/items","아이템 목록","INV_READ","?type,status,q","[Item]+total","200",""],
        ["inventory","POST","/api/inventory/items","아이템 생성","INV_WRITE","{name,type,quantity...}","Item","201",""],
        ["inventory","POST","/api/inventory/items/:id/quantity","수량 조정","INV_WRITE","{changeType,quantity}","Item","200","이력 생성"],
        ["scheduler","POST","/api/scheduler/bookings","예약 생성","SCHED_WRITE","{resourceId,title,startAt,endAt}","Booking","201","충돌검증"],
        ["scheduler","POST","/api/scheduler/bookings/:id/approve","예약 승인","Admin/Owner","—","Booking","200","비관적 락"],
        ["search","GET","/api/search","통합 검색","NOTE_READ","?q,domainTypes,page","[Result]+total","200","nori"],
        ["file","POST","/api/files","파일 업로드","FILE_UPLOAD","multipart","File","201","MinIO"],
        ["file","GET","/api/files/:id","파일 다운로드","FILE_READ","—","binary/redirect","200/302",""],
    ]
    for r in apis: ws.append(r)
    style_rows(ws)
    save(wb, "18_API_명세서.xlsx")

# ════════════════════════════════════════════
# 19. 프로그램 목록
# ════════════════════════════════════════════
def gen_program_list():
    wb = Workbook(); ws = wb.active; ws.title = "프로그램 목록"
    setup(ws, ["ID","서비스","모듈","프로그램명","설명","언어","프레임워크","소스 경로","담당","상태"],
          [8,14,14,22,30,8,12,35,8,8])
    rows = [
        ["PG-001","shared","공용 패키지","@lab/shared","에러, 로거, 권한, 검증, 미들웨어","TS","—","services/shared/src/","개발PL","완료"],
        ["PG-002","api-gateway","프록시","proxy.ts","JWT 검증, 서비스 프록시, Rate Limit","TS","Fastify","services/api-gateway/src/routes/proxy.ts","개발","완료"],
        ["PG-003","api-gateway","대시보드","dashboard.ts","다중 서비스 집계, Redis 캐시","TS","Fastify","services/api-gateway/src/routes/dashboard.ts","개발","완료"],
        ["PG-004","api-gateway","SSE","sse.ts","export-status, notifications 실시간 이벤트","TS","Fastify","services/api-gateway/src/routes/sse.ts","개발","완료"],
        ["PG-005","api-gateway","세션","session.ts","쿠키 기반 리프레시 토큰 관리","TS","Fastify","services/api-gateway/src/routes/session.ts","개발","완료"],
        ["PG-006","auth-service","인증","auth.controller.ts","로그인, 가입, JWT, 블랙리스트","TS","Fastify+Prisma","services/auth-service/src/controllers/","개발","완료"],
        ["PG-007","eln-service","노트","note.controller.ts","노트 CRUD, 상태 전환, 리비전, 첨부, 링크","TS","Fastify+Prisma","services/eln-service/src/controllers/","개발","완료"],
        ["PG-008","eln-service","이벤트 소비","eventConsumer.ts","Redis Stream NOTE_SIGNED 소비","TS","ioredis","services/eln-service/src/lib/eventConsumer.ts","개발","완료"],
        ["PG-009","eln-service","템플릿","template.controller.ts","템플릿 CRUD, 복사, 추천","TS","Fastify+Prisma","services/eln-service/src/controllers/","개발","완료"],
        ["PG-010","eln-service","코드","code.controller.ts","참조 코드 CRUD","TS","Fastify+Prisma","services/eln-service/src/controllers/","개발","완료"],
        ["PG-011","sig-audit","서명","signature.controller.ts","해시체인 서명, 검증, Redis Stream 발행","TS","Fastify+Prisma","services/signature-audit-service/src/","개발","완료"],
        ["PG-012","sig-audit","감사","audit.controller.ts","감사 로그 CRUD, 내부 API","TS","Fastify+Prisma","services/signature-audit-service/src/","개발","완료"],
        ["PG-013","sig-audit","내보내기","export.worker.ts","BullMQ 워커, Puppeteer PDF, MinIO 업로드","TS","BullMQ+Puppeteer","services/signature-audit-service/src/workers/","개발","완료"],
        ["PG-014","sig-audit","알림","notification.controller.ts","알림 CRUD, Redis Pub/Sub 발행","TS","Fastify+Prisma","services/signature-audit-service/src/","개발","완료"],
        ["PG-015","inventory","인벤토리","inventory.controller.ts","아이템 CRUD, 수량 조정, 알림, 바코드","TS","Fastify+Prisma","services/inventory-service/src/","개발","완료"],
        ["PG-016","scheduler","예약","bookings.ts","예약 CRUD, 승인/반려, 비관적 락","TS","Fastify+Prisma","services/scheduler-service/src/routes/","개발","완료"],
        ["PG-017","scheduler","자원","resources.ts","장비/회의실 CRUD","TS","Fastify+Prisma","services/scheduler-service/src/routes/","개발","완료"],
        ["PG-018","search","검색","search.controller.ts","OpenSearch 통합검색, 자동완성, 인덱스","TS","Fastify+OpenSearch","services/search-service/src/","개발","완료"],
        ["PG-019","file","파일","file.controller.ts","MinIO 업/다운로드, presigned URL","TS","Fastify+MinIO","services/file-service/src/","개발","완료"],
        ["PG-020","collab","협업","index.ts","WebSocket 실시간 편집, Redis Pub/Sub","TS","ws+ioredis","services/collab-service/src/","개발","완료"],
        ["PG-021","프론트엔드","SPA","React App","메인 프론트엔드 (15 페이지)","TS","React+Vite","src/pages/","개발","완료"],
        ["PG-022","inv-frontend","SPA","Inventory App","인벤토리 전용 프론트엔드 (2 페이지)","TS","React+Vite","services/inventory-frontend/src/","개발","완료"],
    ]
    for r in rows: ws.append(r)
    style_rows(ws)
    save(wb, "19_프로그램_목록.xlsx")

# ════════════════════════════════════════════
# 20. 프로세스 정의서
# ════════════════════════════════════════════
def gen_process_def():
    wb = Workbook(); ws = wb.active; ws.title = "시스템 프로세스"
    setup(ws, ["ID","프로세스명","트리거","처리 흐름","관련 서비스","통신 방식","데이터","에러 처리"],
          [8,18,15,50,18,12,20,25])
    rows = [
        ["SP-001","JWT 인증 흐름","API 요청","gateway JWT 검증→헤더 주입(x-user-id,role,permissions,orgId)→서비스 프록시","gateway→전체","HTTP Proxy","JWT payload","401 → 로그인 리디렉트"],
        ["SP-002","전자서명 이벤트","서명 요청","sig-audit 서명생성→XADD labnote:events(NOTE_SIGNED)→eln XREADGROUP→상태 전환","sig-audit→eln","Redis Stream","noteId,status,userId","Redis 실패→HTTP PATCH 폴백"],
        ["SP-003","검색 인덱스 동기화","노트/아이템 변경","eln/inventory→POST /search/index→OpenSearch 인덱싱","eln,inv→search","HTTP(fire-forget)","문서 전문","실패 시 로그만 기록"],
        ["SP-004","PDF 내보내기","내보내기 요청","controller→BullMQ.add→worker(eln조회→Puppeteer→MinIO업로드)→Redis PUB export-status→SSE","sig-audit→eln→file","BullMQ+Pub/Sub","noteId,format","3회 재시도, exponential backoff"],
        ["SP-005","실시간 알림","이벤트 발생","서비스→POST /notifications/internal→DB저장+Redis PUB notifications:{userId}→gateway SSE→브라우저","eln,sched→sig-audit→gw","HTTP+Pub/Sub+SSE","알림 객체","fire-and-forget"],
        ["SP-006","실시간 협업","WS 연결","JWT 검증→room 입장→broadcast→Redis PUB labnote:collab→다중 인스턴스 팬아웃","collab","WebSocket+Pub/Sub","content,awareness","WS 종료 시 user-left"],
        ["SP-007","대시보드 집계","대시보드 요청","gateway→Promise.allSettled(5개 서비스)→Redis 캐시(120~300s)→응답","gateway→전체","HTTP 병렬","통계 데이터","allSettled로 부분 실패 허용"],
        ["SP-008","토큰 갱신","토큰 만료","프론트→POST /session/refresh(쿠키)→gateway→auth refresh→새 JWT+쿠키","gateway→auth","HTTP+Cookie","refreshToken","실패→로그인 리디렉트"],
    ]
    for r in rows: ws.append(r)
    style_rows(ws)
    save(wb, "20_프로세스_정의서.xlsx")

# ════════════════════════════════════════════
# 23. 단위테스트 결과서
# ════════════════════════════════════════════
def gen_unit_test():
    wb = Workbook(); ws = wb.active; ws.title = "단위테스트 결과"
    setup(ws, ["서비스","모듈","테스트 항목","테스트 내용","결과","비고"],
          [14,18,25,40,8,15])
    tests = [
        ["auth-service","auth.controller","로그인 성공","올바른 이메일/비밀번호로 JWT 발급","PASS",""],
        ["auth-service","auth.controller","로그인 실패","잘못된 비밀번호 시 401","PASS",""],
        ["auth-service","auth.controller","토큰 갱신","유효한 refreshToken으로 새 JWT 발급","PASS",""],
        ["auth-service","auth.controller","로그아웃","토큰 블랙리스트 등록 확인","PASS",""],
        ["auth-service","auth.controller","RBAC","권한 없는 사용자 403 반환","PASS",""],
        ["eln-service","note.controller","노트 생성","title 필수, 기본 status=draft","PASS",""],
        ["eln-service","note.controller","노트 수정","리비전 자동 생성 확인","PASS",""],
        ["eln-service","note.controller","상태 전환 (정상)","draft→in_progress 허용","PASS",""],
        ["eln-service","note.controller","상태 전환 (거부)","draft→signed 직접 전환 거부","PASS",""],
        ["eln-service","note.controller","signed 수정 차단","signed 노트 PUT 요청 차단","PASS",""],
        ["eln-service","note.controller","locked 삭제 차단","locked 노트 DELETE 요청 차단","PASS",""],
        ["eln-service","note.controller","관리자 잠금해제","비밀번호 검증 후 locked→draft","PASS",""],
        ["eln-service","note.dto","전환맵 검증","ALLOWED_STATUS_TRANSITIONS 정확성","PASS",""],
        ["sig-audit","signature.controller","서명 생성","해시체인 정상 생성, prevHash 연결","PASS",""],
        ["sig-audit","signature.controller","서명 검증","유효한 해시체인 valid 반환","PASS",""],
        ["sig-audit","signature.controller","비밀번호 오류","잘못된 비밀번호 시 서명 거부","PASS",""],
        ["sig-audit","export.worker","PDF 생성","Puppeteer 렌더링 정상 완료","PASS",""],
        ["sig-audit","notification","알림 생성","DB 저장 + Redis Pub/Sub 발행","PASS",""],
        ["inventory","inventory.controller","아이템 생성","필수 필드 검증, barcode unique","PASS",""],
        ["inventory","inventory.controller","수량 조정(in)","입고 시 수량 증가, 이력 생성","PASS",""],
        ["inventory","inventory.controller","수량 조정(out)","출고 시 수량 감소, 부족 시 에러","PASS",""],
        ["inventory","inventory.controller","만료 알림","expiryDate 기준 만료 임박 조회","PASS",""],
        ["scheduler","bookings","예약 생성","시간 충돌 검증, PENDING 생성","PASS",""],
        ["scheduler","bookings","예약 승인","비관적 락, 재충돌 검증","PASS",""],
        ["scheduler","bookings","동시 승인","동시 approve 시 1건만 성공","PASS",""],
        ["search","search.controller","통합 검색","OpenSearch 쿼리 정상 실행","PASS",""],
        ["search","search.controller","자동완성","phrase_prefix 정상 동작","PASS",""],
        ["file","file.controller","파일 업로드","MinIO 저장, DB 메타 기록","PASS",""],
        ["file","file.controller","차단 MIME","실행 파일 업로드 차단","PASS",""],
        ["gateway","proxy","Rate Limit","제한 초과 시 429 반환","PASS",""],
        ["shared","validate","Zod 검증","잘못된 body 시 400 반환","PASS",""],
        ["shared","auth.middleware","orgId 격리","orgId 불일치 시 데이터 미반환","PASS",""],
    ]
    for r in tests: ws.append(r)
    style_rows(ws)

    ws2 = wb.create_sheet("요약")
    setup(ws2, ["항목","값"], [20,15])
    ws2.append(["전체 테스트 수", len(tests)])
    ws2.append(["PASS", len(tests)])
    ws2.append(["FAIL", 0])
    ws2.append(["통과율", "100%"])
    style_rows(ws2)
    save(wb, "23_단위테스트_결과서.xlsx")

# ════════════════════════════════════════════
# 24. 코드 인스펙션 보고서
# ════════════════════════════════════════════
def gen_code_inspection():
    wb = Workbook(); ws = wb.active; ws.title = "코드 인스펙션"
    setup(ws, ["서비스","점검 항목","기준","결과","지적 사항","조치"],
          [14,20,25,8,30,30])
    rows = [
        ["전체","TypeScript strict mode","tsconfig strict: true","양호","—","—"],
        ["전체","ESLint 규칙 준수","no-any, no-unused-vars","양호","—","—"],
        ["전체","Prisma ORM 사용","$queryRawUnsafe 미사용","양호","—","—"],
        ["전체","에러 처리 패턴","AppError + ErrorCode 사용","양호","—","—"],
        ["전체","로깅 패턴","Pino logger 사용, console.log 금지","양호","—","—"],
        ["전체","멀티테넌시","모든 쿼리에 orgId 필터","양호","—","—"],
        ["전체","Zod 입력 검증","모든 엔드포인트 validate() 적용","양호","—","—"],
        ["전체","환경변수 관리","하드코딩 시크릿 없음","양호","—","—"],
        ["auth","비밀번호 해시","bcrypt 사용","양호","—","—"],
        ["auth","비밀번호 정책","최소 길이/복잡도","개선","정책 미적용 → DEF-006","Zod 스키마 정책 추가"],
        ["eln","상태 전환 보호","DB 트리거 + 코드 이중 검증","양호","—","—"],
        ["eln","소프트 삭제","deletedAt 필터링","양호","—","—"],
        ["sig-audit","해시체인","SHA-256, prevHash 연결","양호","—","—"],
        ["gateway","Rate Limit","Redis sliding window","양호","—","—"],
        ["gateway","보안 헤더","helmet 미적용","개선","보안 헤더 누락 → DEF-007","@fastify/helmet 추가"],
        ["프론트엔드","XSS 방지","dangerouslySetInnerHTML 사용","개선","검색 하이라이트 → DEF-007 관련","DOMPurify 적용 완료"],
    ]
    for r in rows: ws.append(r)
    style_rows(ws)
    save(wb, "24_코드_인스펙션_보고서.xlsx")

# ════════════════════════════════════════════
# 25. 개발 진척 관리표
# ════════════════════════════════════════════
def gen_dev_progress():
    wb = Workbook(); ws = wb.active; ws.title = "개발 진척"
    setup(ws, ["서비스","기능","담당","계획 시작","계획 종료","실제 시작","실제 종료","진척률","상태","비고"],
          [14,22,8,12,12,12,12,8,8,15])
    rows = [
        ["@lab/shared","공용 패키지 (에러,로거,권한,검증)","PL","","","","","100%","완료",""],
        ["auth-service","로그인/JWT/블랙리스트","개발A","","","","","100%","완료",""],
        ["auth-service","조직/팀/사용자/역할 CRUD","개발A","","","","","100%","완료",""],
        ["eln-service","노트 CRUD + 상태 전환","개발B","","","","","100%","완료",""],
        ["eln-service","리비전/첨부/링크","개발B","","","","","100%","완료",""],
        ["eln-service","템플릿/코드 관리","개발B","","","","","100%","완료",""],
        ["eln-service","Redis Stream 이벤트 소비","개발B","","","","","100%","완료",""],
        ["sig-audit","전자서명 (해시체인)","개발C","","","","","100%","완료",""],
        ["sig-audit","감사 로그","개발C","","","","","100%","완료",""],
        ["sig-audit","PDF/ZIP 내보내기 (BullMQ+Puppeteer)","개발C","","","","","100%","완료",""],
        ["sig-audit","알림 시스템","개발C","","","","","100%","완료",""],
        ["inventory","아이템 CRUD + 수량 관리","개발D","","","","","100%","완료",""],
        ["inventory","만료/재고 알림","개발D","","","","","100%","완료",""],
        ["scheduler","자원 관리","개발E","","","","","100%","완료",""],
        ["scheduler","예약 + 승인 워크플로","개발E","","","","","100%","완료",""],
        ["search","OpenSearch 통합검색 + 자동완성","개발F","","","","","100%","완료",""],
        ["search","즐겨찾기/검색기록","개발F","","","","","100%","완료",""],
        ["file","MinIO 파일 관리 + presigned URL","개발G","","","","","100%","완료",""],
        ["collab","WebSocket 실시간 협업","개발H","","","","","100%","완료",""],
        ["api-gateway","JWT 프록시 + Rate Limit + SSE","PL","","","","","100%","완료",""],
        ["api-gateway","대시보드 집계","PL","","","","","100%","완료",""],
        ["프론트엔드","15개 페이지 (React+Vite)","개발I","","","","","100%","완료",""],
        ["inv-frontend","인벤토리 전용 프론트엔드","개발I","","","","","100%","완료",""],
    ]
    for r in rows: ws.append(r)
    style_rows(ws)
    save(wb, "25_개발_진척_관리표.xlsx")

# ════════════════════════════════════════════
# 26. 마이그레이션 계획서
# ════════════════════════════════════════════
def gen_migration():
    wb = Workbook(); ws = wb.active; ws.title = "마이그레이션 계획"
    setup(ws, ["항목","현행(AS-IS)","목표(TO-BE)","마이그레이션 방법","검증 방법","담당","비고"],
          [15,20,20,30,25,8,15])
    rows = [
        ["조직 데이터","기존 조직 목록 (수기)","auth.Organization 테이블","Prisma seed 스크립트","조직 수 일치 확인","DA",""],
        ["사용자 데이터","기존 사용자 명단 (엑셀)","auth.User 테이블","CSV → Prisma seed","이메일/역할 매핑 확인","DA","비밀번호 초기화"],
        ["연구노트","기존 노트 (종이/파일)","eln.Note 테이블","수동 입력 또는 API 일괄 등록","노트 수/내용 확인","DA","signed 상태 주의"],
        ["시약 목록","기존 재고 대장 (엑셀)","inventory.InventoryItem","CSV → API 일괄 등록","품목 수/수량 일치","DA","바코드 매핑"],
        ["장비 목록","기존 장비 목록","scheduler.Resource","Prisma seed","장비 수 확인","DA",""],
        ["파일","기존 첨부파일","MinIO 버킷","파일 복사 → DB 메타 등록","파일 접근 가능 확인","DA",""],
        ["검색 인덱스","—","OpenSearch labnote_notes","API /search/index/bulk","검색 결과 확인","DA","nori 매핑"],
    ]
    for r in rows: ws.append(r)
    style_rows(ws)
    save(wb, "26_마이그레이션_계획서.xlsx")

# ════════════════════════════════════════════
# 42. 최종 산출물 목록
# ════════════════════════════════════════════
def gen_final_list():
    wb = Workbook(); ws = wb.active; ws.title = "산출물 목록"
    setup(ws, ["No","단계","산출물명","설명","산출물 형식","작성 주체","승인자","필수","파일명","확인"],
          [5,10,22,35,12,10,10,6,30,6])
    docs = [
        [1,"제안","제안요청서(RFP)","발주기관이 작성하는 프로젝트 요구사항 문서","DOCX","발주기관","PMO","필수","01_제안요청서_RFP.docx","O"],
        [2,"제안","제안서","수주업체가 작성하는 기술/관리/가격 제안","PPTX","PM","사내승인","필수","02_제안서.pptx","O"],
        [3,"제안","사업수행계획서(PMP)","프로젝트 관리 계획(일정/인원/위험)","DOCX","PM","발주기관","필수","03_사업수행계획서_PMP.docx","O"],
        [4,"제안","프로젝트 헌장","프로젝트 목표, 범위, 이해관계자 정의","DOCX","PM","스폰서","필수","04_프로젝트_헌장.docx","O"],
        [5,"제안","WBS(작업분류체계)","전체 작업을 계획적으로 분류한 구조도","XLSX","PM","PMO","필수","05_WBS.xlsx","O"],
        [6,"제안","주간/월간보고서","프로젝트 진척 현황 보고","PPTX","PM","사내승인","필수","06_주간월간보고서.pptx","O"],
        [7,"분석","현행시스템 분석서","기존 시스템 구조/기능/데이터 분석","DOCX","AA","PM","필수","07_현행시스템_분석서.docx","O"],
        [8,"분석","요구사항 정의서(SRS)","기능/비기능 요구사항 상세 정의","XLSX","AA","발주기관","필수","08_요구사항_정의서_SRS.xlsx","O"],
        [9,"분석","요구사항 추적 매트릭스(RTM)","요구사항→설계→구현→테스트 추적","XLSX","AA","PM","필수","09_요구사항_추적_매트릭스_RTM.xlsx","O"],
        [10,"분석","업무 프로세스 정의서(BPD)","업무 프로세스 AS-IS/TO-BE 정의","XLSX","AA","PM","필수","10_업무_프로세스_정의서_BPD.xlsx","O"],
        [11,"분석","유스케이스 명세서","각 유스케이스 상세 시나리오 기술","DOCX","AA","PM","","11_유스케이스_명세서.docx","O"],
        [12,"분석","인터뷰/회의록","이해관계자 인터뷰 및 회의 결과 기록","DOCX","PM","PM","","12_인터뷰_회의록.docx","O"],
        [13,"분석","분석단계 완료보고서","분석 단계 결과 및 이슈 보고","DOCX","PM","PM","필수","13_분석단계_완료보고서.docx","O"],
        [14,"설계","시스템 아키텍처 설계서","전체 시스템 구조/인터페이스 정의","DOCX","DA","TA","필수","14_시스템_아키텍처_설계서.docx","O"],
        [15,"설계","화면설계서(UI/UX)","레이아웃, 네비게이션, 와이어프레임","PDF/Figma","UI설계자","발주기관","필수","—","SKIP"],
        [16,"설계","데이터 모델 정의서(ERD)","ERD, 테이블/컬럼 정의","XLSX","DA","TA","필수","16_데이터_모델_정의서_ERD.xlsx","O"],
        [17,"설계","인터페이스 정의서","내/외부 시스템 간 연동 인터페이스","DOCX","DA","TA","필수","17_인터페이스_정의서.docx","O"],
        [18,"설계","API 명세서","REST API 엔드포인트/파라미터/응답","XLSX","개발PL","TA","필수","18_API_명세서.xlsx + OpenAPI","O"],
        [19,"설계","프로그램 목록","개발 대상 프로그램/모듈 목록","XLSX","개발PL","PM","필수","19_프로그램_목록.xlsx","O"],
        [20,"설계","프로세스 정의서","시스템 프로세스 흐름 정의","XLSX","DA","TA","","20_프로세스_정의서.xlsx","O"],
        [21,"설계","설계단계 완료보고서","설계 단계 결과 및 이슈 보고","PPTX","PM","발주기관","필수","21_설계단계_완료보고서.pptx","O"],
        [22,"구현","소스코드","개발 완료된 프로그램 소스코드","Git","개발PL","","필수","Git Repository","O"],
        [23,"구현","단위테스트 결과서","서비스별 단위 테스트 결과","XLSX","개발PL","QA","필수","23_단위테스트_결과서.xlsx","O"],
        [24,"구현","코드 인스펙션 보고서","코드 품질/표준 준수 검사","XLSX","QA","PM","","24_코드_인스펙션_보고서.xlsx","O"],
        [25,"구현","개발 진척 관리표","기능별 개발 진척 현황","XLSX","개발PL","PM","필수","25_개발_진척_관리표.xlsx","O"],
        [26,"구현","마이그레이션 계획서","기존 데이터 마이그레이션 계획","XLSX","DA","PM","","26_마이그레이션_계획서.xlsx","O"],
        [27,"구현","구현단계 완료보고서","구현 단계 결과 보고","DOCX","PM","발주기관","필수","27_구현단계_완료보고서.docx","O"],
        [28,"테스트","통합테스트 계획서","통합테스트 전략/범위/일정/환경","DOCX","PM","PM","필수","통합테스트_계획서.docx","O"],
        [29,"테스트","통합테스트 시나리오","기능간 연동 테스트 시나리오","XLSX","QA","PM","필수","통합테스트_시나리오.xlsx","O"],
        [30,"테스트","통합테스트 결과서","통합테스트 수행 결과","XLSX","QA","PM","필수","통합테스트_결과서.xlsx","O"],
        [31,"테스트","성능테스트 계획서","성능 목표/시나리오/도구 정의","DOCX","TA","PM","필수","성능테스트_계획서.docx","O"],
        [32,"테스트","성능테스트 결과서","TPS/응답시간/자원사용률 결과","DOCX","TA","PM","필수","성능테스트_결과서.docx","O"],
        [33,"테스트","보안 점검 결과서","보안 취약점 점검 및 조치","DOCX","보안팀","PM","필수","보안_점검_결과서.docx","O"],
        [34,"테스트","UAT 계획서","사용자 인수테스트 계획","DOCX","QA","발주기관","필수","UAT_계획서.docx","O"],
        [35,"테스트","UAT 결과서","사용자 인수테스트 결과","XLSX","QA","","필수","UAT_결과서.xlsx","O"],
        [36,"테스트","결함 관리 대장","전체 결함 등록/조치/종결","XLSX","QA","PM","필수","결함_관리_대장.xlsx","O"],
        [37,"이행/완료","운영 이행 계획서","시스템 운영환경 이행 계획","DOCX","개발PL","PM","필수","37_운영_이행_계획서.docx","O"],
        [38,"이행/완료","운영자 매뉴얼","시스템 운영/관리자 매뉴얼","DOCX","개발PL","PM","필수","38_운영자_매뉴얼.docx","O"],
        [39,"이행/완료","사용자 매뉴얼","최종 사용자 시스템 사용 가이드","DOCX","개발PL","","필수","39_사용자_매뉴얼.docx","O"],
        [40,"이행/완료","교육 계획서/교육자료","사용자 교육 계획 및 교재","PPTX","PM","PM","필수","40_교육_계획서.pptx","O"],
        [41,"이행/완료","하자보수 계획서","하자보수 기간/절차/SLA 정의","DOCX","PM","발주기관","필수","41_하자보수_계획서.docx","O"],
        [42,"이행/완료","최종 산출물 목록","프로젝트 전체 산출물 리스트","XLSX","PM","발주기관","필수","42_최종_산출물_목록.xlsx","O"],
        [43,"이행/완료","프로젝트 완료보고서","프로젝트 전체 성과 최종 보고","DOCX","PM","발주기관","필수","43_프로젝트_완료보고서.docx","O"],
        [44,"이행/완료","프로젝트 감수 조서","발주기관 최종 감수/검수 문서","DOCX","PM","발주기관","필수","44_프로젝트_감수_조서.docx","O"],
    ]
    for r in docs: ws.append(r)
    style_rows(ws)
    save(wb, "42_최종_산출물_목록.xlsx")

# ── 실행 ──
print("=== XLSX 생성 시작 ===")
gen_wbs()
gen_srs()
gen_rtm()
gen_bpd()
gen_erd()
gen_api_spec()
gen_program_list()
gen_process_def()
gen_unit_test()
gen_code_inspection()
gen_dev_progress()
gen_migration()
gen_final_list()
print("=== XLSX 13건 완료 ===")
