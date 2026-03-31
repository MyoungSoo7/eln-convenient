"""시퀀스 다이어그램을 이미지로 그려서 엑셀 파일에 삽입하는 스크립트"""
import os
import tempfile
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch, FancyArrowPatch
import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 한글 폰트 설정 ──
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

# ── 색상 팔레트 ──
ACTOR_COLOR = '#FFD54F'
SERVICE_COLOR = '#42A5F5'
DB_COLOR = '#66BB6A'
INFRA_COLOR = '#EF5350'
ARROW_COLOR = '#37474F'
ERROR_COLOR = '#E53935'
ASYNC_COLOR = '#FF9800'
NOTE_BG = '#FFF9C4'
ALT_BG = '#FFEBEE'
TX_BG = '#E3F2FD'
LOOP_BG = '#F3E5F5'
LIFELINE_COLOR = '#B0BEC5'

PARTICIPANT_COLORS = {
    'actor': ACTOR_COLOR,
    'frontend': '#81D4FA',
    'gateway': '#7E57C2',
    'auth': '#26A69A',
    'eln': '#5C6BC0',
    'signature': '#EC407A',
    'inventory': '#8D6E63',
    'scheduler': '#78909C',
    'search': '#29B6F6',
    'file': '#FFA726',
    'collab': '#AB47BC',
    'db': DB_COLOR,
    'redis': INFRA_COLOR,
    'opensearch': '#1565C0',
    'minio': '#C62828',
    'bullmq': '#F44336',
    'puppeteer': '#4CAF50',
    'worker': '#FF7043',
}


def get_color(name):
    nl = name.lower()
    for key, color in PARTICIPANT_COLORS.items():
        if key in nl:
            return color
    return '#90A4AE'


def draw_sequence_diagram(title, participants, messages, fig_width=20, notes=None, regions=None):
    """
    시퀀스 다이어그램을 matplotlib로 그립니다.

    participants: [(name, short_label)] - 참여자 목록
    messages: [(from_idx, to_idx, label, style)] - 메시지 목록
        style: 'normal', 'error', 'async', 'self', 'return'
    notes: [(over_idx, text, bg_color)] - 노트
    regions: [(start_msg_idx, end_msg_idx, label, bg_color)] - alt/loop/tx 영역
    """
    if notes is None:
        notes = []
    if regions is None:
        regions = []

    n_participants = len(participants)
    n_messages = len(messages)

    # 레이아웃 계산
    col_spacing = fig_width / (n_participants + 1)
    row_height = 0.7
    header_y = n_messages * row_height + 3
    fig_height = header_y + 2

    fig, ax = plt.subplots(1, 1, figsize=(fig_width, fig_height))
    ax.set_xlim(-0.5, fig_width + 0.5)
    ax.set_ylim(-1, fig_height + 0.5)
    ax.axis('off')

    # 타이틀
    ax.text(fig_width / 2, fig_height - 0.2, title,
            ha='center', va='top', fontsize=16, fontweight='bold',
            color='#1A237E')

    # 참여자 위치 계산
    px = [(i + 1) * col_spacing for i in range(n_participants)]
    top_y = fig_height - 1.5
    bottom_y = 0.5

    # 참여자 박스 + 라이프라인
    for i, (name, label) in enumerate(participants):
        x = px[i]
        color = get_color(name)

        # 상단 박스
        box_w = col_spacing * 0.85
        box = FancyBboxPatch((x - box_w/2, top_y - 0.3), box_w, 0.7,
                             boxstyle="round,pad=0.1",
                             facecolor=color, edgecolor='#37474F',
                             linewidth=1.5, alpha=0.9)
        ax.add_patch(box)
        # 텍스트 줄바꿈
        display_label = label.replace('\\n', '\n')
        ax.text(x, top_y + 0.05, display_label,
                ha='center', va='center', fontsize=7.5, fontweight='bold',
                color='white' if color not in [ACTOR_COLOR, '#81D4FA', NOTE_BG] else '#212121')

        # 라이프라인
        ax.plot([x, x], [top_y - 0.3, bottom_y], color=LIFELINE_COLOR,
                linewidth=1, linestyle='--', alpha=0.6, zorder=0)

        # 하단 박스
        box2 = FancyBboxPatch((x - box_w/2, bottom_y - 0.5), box_w, 0.5,
                              boxstyle="round,pad=0.05",
                              facecolor=color, edgecolor='#37474F',
                              linewidth=1, alpha=0.7)
        ax.add_patch(box2)
        ax.text(x, bottom_y - 0.25, label.split('\\n')[0],
                ha='center', va='center', fontsize=6.5, color='white' if color not in [ACTOR_COLOR, '#81D4FA'] else '#212121')

    # 영역(alt/loop/tx) 그리기 — 메시지 뒤에 그리면 가려지므로 먼저
    for start_idx, end_idx, region_label, bg_color in regions:
        y1 = top_y - 1.2 - start_idx * row_height + 0.25
        y2 = top_y - 1.2 - end_idx * row_height - 0.25
        rect = plt.Rectangle((0.3, y2), fig_width - 0.6, y1 - y2,
                              facecolor=bg_color, edgecolor='#78909C',
                              linewidth=1.5, linestyle='--', alpha=0.3, zorder=0)
        ax.add_patch(rect)
        # 라벨 배지
        badge = FancyBboxPatch((0.3, y1 - 0.25), len(region_label) * 0.13 + 0.4, 0.25,
                               boxstyle="round,pad=0.05",
                               facecolor='#78909C', edgecolor='none', alpha=0.8)
        ax.add_patch(badge)
        ax.text(0.5, y1 - 0.12, region_label, fontsize=6.5, fontweight='bold', color='white')

    # 메시지 화살표
    for msg_idx, (from_i, to_i, label, style) in enumerate(messages):
        y = top_y - 1.2 - msg_idx * row_height
        x1 = px[from_i]
        x2 = px[to_i]

        if style == 'error':
            color = ERROR_COLOR
            ls = '-'
            lw = 1.5
        elif style == 'async':
            color = ASYNC_COLOR
            ls = '--'
            lw = 1.2
        elif style == 'return':
            color = '#78909C'
            ls = '--'
            lw = 1
        elif style == 'self':
            color = ARROW_COLOR
            ls = '-'
            lw = 1.2
        else:
            color = ARROW_COLOR
            ls = '-'
            lw = 1.2

        if from_i == to_i:
            # 자기 자신에게 보내는 메시지
            loop_w = col_spacing * 0.3
            ax.annotate('', xy=(x1 + 0.05, y - 0.15),
                        xytext=(x1 + loop_w, y - 0.15),
                        arrowprops=dict(arrowstyle='->', color=color, lw=lw, ls=ls))
            ax.plot([x1, x1 + loop_w], [y, y], color=color, lw=lw, ls=ls)
            ax.plot([x1 + loop_w, x1 + loop_w], [y, y - 0.15], color=color, lw=lw, ls=ls)
            ax.text(x1 + loop_w + 0.1, y - 0.05, label,
                    ha='left', va='center', fontsize=6, color=color,
                    bbox=dict(boxstyle='round,pad=0.15', facecolor='white', edgecolor='none', alpha=0.8))
        else:
            # 일반 화살표
            ax.annotate('', xy=(x2, y), xytext=(x1, y),
                        arrowprops=dict(arrowstyle='->', color=color, lw=lw,
                                        linestyle=ls, connectionstyle='arc3,rad=0'))

            # 라벨 위치
            mid_x = (x1 + x2) / 2
            offset_y = 0.12
            txt_color = color if style in ('error', 'async') else '#212121'

            # 긴 텍스트 줄바꿈
            display = label
            if len(label) > 40:
                mid = len(label) // 2
                # 공백 기준 줄바꿈
                space_pos = label.rfind(' ', 0, mid + 10)
                if space_pos > 10:
                    display = label[:space_pos] + '\n' + label[space_pos+1:]

            ax.text(mid_x, y + offset_y, display,
                    ha='center', va='bottom', fontsize=5.8, color=txt_color,
                    bbox=dict(boxstyle='round,pad=0.1', facecolor='white', edgecolor='none', alpha=0.85))

    # 노트 그리기
    for over_idx, note_text, bg in notes:
        # 해당 메시지 인덱스 위치에 노트
        if isinstance(over_idx, tuple):
            x_pos = (px[over_idx[0]] + px[over_idx[1]]) / 2
            y_pos = top_y - 0.6
        else:
            x_pos = px[over_idx]
            y_pos = top_y - 0.6
        note_box = FancyBboxPatch((x_pos - 1.5, y_pos - 0.2), 3, 0.4,
                                   boxstyle="round,pad=0.1",
                                   facecolor=bg, edgecolor='#FBC02D',
                                   linewidth=1, alpha=0.9)
        ax.add_patch(note_box)
        ax.text(x_pos, y_pos, note_text, ha='center', va='center',
                fontsize=6, style='italic', color='#5D4037')

    plt.tight_layout(pad=0.5)
    return fig


# ── 시퀀스 다이어그램 데이터 정의 ──

diagrams = []

# 1-1. 로그인
diagrams.append({
    'title': '1-1. 로그인',
    'sheet': '1-1. 로그인',
    'participants': [
        ('actor', '사용자'),
        ('frontend', 'Frontend\\n(React)'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('auth', 'Auth Service\\n(:8001)'),
        ('db', 'PostgreSQL\\n(auth)'),
        ('redis', 'Redis'),
    ],
    'messages': [
        (0, 1, '이메일/비밀번호 입력 후 로그인 클릭', 'normal'),
        (1, 2, 'POST /api/auth/login {email, password}', 'normal'),
        (2, 3, '프록시 전달 (공개 경로 — JWT 생략)', 'normal'),
        (3, 4, 'SELECT user WHERE email (+ role, teams 조인)', 'normal'),
        (4, 3, '사용자 정보 반환', 'return'),
        (3, 3, 'bcrypt.compare(password, hash)', 'self'),
        (3, 1, '[실패] 401 이메일 또는 비밀번호 오류', 'error'),
        (3, 3, 'Access Token 생성 (15분)', 'self'),
        (3, 3, 'Refresh Token 생성 (8시간)', 'self'),
        (3, 2, '{ok:true, data:{token, refreshToken, user}}', 'return'),
        (2, 1, '응답 전달', 'return'),
        (1, 1, 'Access Token → 메모리 저장', 'self'),
        (1, 2, 'POST /api/auth/session {refreshToken}', 'normal'),
        (2, 2, 'Set-Cookie: labnote_rt (HttpOnly, Secure)', 'self'),
        (2, 1, '200 OK', 'return'),
        (1, 0, '대시보드로 이동', 'normal'),
    ],
})

# 1-2. 토큰 갱신
diagrams.append({
    'title': '1-2. 토큰 갱신 (자동)',
    'sheet': '1-2. 토큰 갱신',
    'participants': [
        ('frontend', 'Frontend\\n(React)'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('auth', 'Auth Service\\n(:8001)'),
        ('redis', 'Redis'),
    ],
    'messages': [
        (0, 0, 'Access Token 만료 감지 (15분)', 'self'),
        (0, 1, 'POST /api/auth/session/refresh (Cookie: labnote_rt)', 'normal'),
        (1, 1, '쿠키에서 Refresh Token 추출', 'self'),
        (1, 2, 'POST /api/auth/refresh (x-user-id 헤더)', 'normal'),
        (2, 2, 'Refresh Token 서명 검증', 'self'),
        (2, 3, 'GET blacklist:user:{userId}', 'normal'),
        (2, 0, '[무효화] 401 "권한이 변경되었습니다"', 'error'),
        (2, 2, '새 Access Token (15분) 발급', 'self'),
        (2, 2, '새 Refresh Token (8시간) 발급', 'self'),
        (2, 1, '{token, refreshToken}', 'return'),
        (1, 1, 'Set-Cookie: labnote_rt 교체', 'self'),
        (1, 0, '{ok:true, data:{token}}', 'return'),
        (0, 0, '새 Access Token → 메모리 교체', 'self'),
    ],
})

# 1-3. 로그아웃
diagrams.append({
    'title': '1-3. 로그아웃',
    'sheet': '1-3. 로그아웃',
    'participants': [
        ('actor', '사용자'),
        ('frontend', 'Frontend\\n(React)'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('auth', 'Auth Service\\n(:8001)'),
        ('redis', 'Redis'),
    ],
    'messages': [
        (0, 1, '로그아웃 클릭', 'normal'),
        (1, 2, 'POST /api/auth/logout (Bearer {accessToken})', 'normal'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 3, '토큰에서 만료시각(exp) 추출', 'self'),
        (3, 4, 'SET blacklist:{token} "1" TTL=남은 만료 시간', 'normal'),
        (3, 2, '{ok:true}', 'return'),
        (1, 2, 'DELETE /api/auth/session', 'normal'),
        (2, 2, 'Set-Cookie: labnote_rt=; Max-Age=0', 'self'),
        (2, 1, '200 OK', 'return'),
        (1, 1, '메모리의 Access Token 삭제', 'self'),
        (1, 0, '로그인 페이지로 이동', 'normal'),
    ],
})

# 1-4. JWT 검증
diagrams.append({
    'title': '1-4. API 요청 시 Gateway JWT 검증',
    'sheet': '1-4. JWT 검증',
    'participants': [
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('redis', 'Redis'),
        ('service', '내부 서비스'),
    ],
    'messages': [
        (0, 1, 'API 요청 (Authorization: Bearer {token})', 'normal'),
        (1, 1, '① 내부 헤더 제거 (스푸핑 방지)', 'self'),
        (1, 1, '② 공개 경로 확인 (/health, /login)', 'self'),
        (1, 1, '③ /internal 경로 차단 (404)', 'self'),
        (1, 2, '④ GET blacklist:{token}', 'normal'),
        (1, 0, '[블랙리스트] 401 "만료된 세션"', 'error'),
        (1, 1, '⑤ JWT 서명 검증 (JWT_SECRET)', 'self'),
        (1, 2, '⑥ GET blacklist:user:{userId}', 'normal'),
        (1, 0, '[무효화] 401 "권한이 변경되었습니다"', 'error'),
        (1, 1, '⑦ 내부 헤더 주입 (x-user-id, role, permissions...)', 'self'),
        (1, 3, '⑧ 프록시 전달 (+ 주입된 헤더)', 'normal'),
        (3, 1, '응답', 'return'),
        (1, 0, '응답 전달', 'return'),
    ],
})

# 2-1. 노트 생성
diagrams.append({
    'title': '2-1. 연구노트 생성',
    'sheet': '2-1. 노트 생성',
    'participants': [
        ('actor', '연구자'),
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('eln', 'ELN Service\\n(:8002)'),
        ('db', 'PostgreSQL\\n(eln)'),
        ('search', 'Search\\n(:8006)'),
        ('signature', 'Sig-Audit\\n(:8003)'),
    ],
    'messages': [
        (0, 1, '새 연구노트 작성 클릭', 'normal'),
        (1, 2, 'POST /api/notes {title, content, type, tags}', 'normal'),
        (2, 2, 'JWT 검증 + 헤더 주입', 'self'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 4, "INSERT note (status:'draft')", 'normal'),
        (4, 3, '생성된 노트', 'return'),
        (3, 4, 'INSERT noteRevision (revision:1)', 'normal'),
        (3, 4, '[opt] UPDATE template useCount+1', 'normal'),
        (3, 5, '[비동기] POST /api/search/index', 'async'),
        (3, 6, '[비동기] 감사로그 note.created', 'async'),
        (3, 2, '201 {ok:true, data:note}', 'return'),
        (2, 1, '응답 전달', 'return'),
        (1, 0, '에디터 화면으로 이동', 'normal'),
    ],
    'regions': [
        (8, 9, 'par (비동기)', '#FFF9C4'),
    ],
})

# 2-2. 노트 수정
diagrams.append({
    'title': '2-2. 연구노트 수정 (저장)',
    'sheet': '2-2. 노트 수정',
    'participants': [
        ('actor', '연구자'),
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('eln', 'ELN Service\\n(:8002)'),
        ('db', 'PostgreSQL\\n(eln)'),
        ('search', 'Search\\n(:8006)'),
        ('signature', 'Sig-Audit\\n(:8003)'),
    ],
    'messages': [
        (0, 1, '노트 내용 수정 후 저장', 'normal'),
        (1, 2, 'PUT /api/notes/:id {title, content, tags}', 'normal'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 4, 'SELECT note WHERE id AND orgId', 'normal'),
        (3, 1, '[없음] 404 NOTE_NOT_FOUND', 'error'),
        (3, 1, '[잠김] 403 NOTE_LOCKED / NOTE_SIGNED', 'error'),
        (3, 3, '소유자 또는 Admin 확인', 'self'),
        (3, 4, 'UPDATE note SET title, content, tags', 'normal'),
        (3, 4, 'INSERT noteRevision (revision:count+1)', 'normal'),
        (3, 5, '[비동기] POST /api/search/index 재인덱싱', 'async'),
        (3, 6, '[비동기] 감사로그 note.updated', 'async'),
        (3, 2, '200 {ok:true, data:updatedNote}', 'return'),
        (2, 1, '응답 전달', 'return'),
        (1, 0, '저장 완료 토스트', 'normal'),
    ],
    'regions': [
        (9, 10, 'par (비동기)', '#FFF9C4'),
    ],
})

# 3. 실시간 협업 편집
diagrams.append({
    'title': '3. 실시간 협업 편집',
    'sheet': '3. 실시간 협업',
    'participants': [
        ('actor', 'Alice\\n(연구자 A)'),
        ('actor', 'Bob\\n(연구자 B)'),
        ('collab', 'Collab Service\\n(:8009 WS)'),
        ('redis', 'Redis\\n(pub/sub)'),
    ],
    'messages': [
        (0, 2, 'WebSocket 연결 ws://.../collab/notes/{noteId}', 'normal'),
        (2, 2, 'JWT 검증 + Room에 Alice 추가 (colorIdx:0)', 'self'),
        (2, 0, "{type:'joined', users:[]}", 'return'),
        (1, 2, 'WebSocket 연결 (같은 노트)', 'normal'),
        (2, 2, 'Room에 Bob 추가 (colorIdx:1)', 'self'),
        (2, 1, "{type:'joined', users:[alice]}", 'return'),
        (2, 0, "{type:'user-joined', userId:'bob'}", 'normal'),
        (2, 3, "PUBLISH labnote:collab (user-joined)", 'normal'),
        (0, 2, "{type:'content-update', content:'실험 결과...'}", 'normal'),
        (2, 1, "content-update 전달 (400ms 디바운스)", 'normal'),
        (2, 3, 'PUBLISH labnote:collab (content-update)', 'normal'),
        (0, 2, "{type:'awareness', cursorLine:15}", 'normal'),
        (2, 1, "awareness 전달 (커서 위치)", 'normal'),
        (1, 2, 'WebSocket 연결 종료', 'normal'),
        (2, 2, 'Room에서 Bob 제거', 'self'),
        (2, 0, "{type:'user-left', userId:'bob'}", 'normal'),
        (2, 3, "PUBLISH labnote:collab (user-left)", 'normal'),
    ],
})

# 4. 전자서명 → 상태 전환
diagrams.append({
    'title': '4. 전자서명 → 상태 전환 (해시체인)',
    'sheet': '4. 전자서명',
    'participants': [
        ('actor', '검토자\\n(Reviewer)'),
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('signature', 'Sig-Audit\\n(:8003)'),
        ('auth', 'Auth\\n(:8001)'),
        ('db', 'PostgreSQL\\n(signature)'),
        ('redis', 'Redis Stream'),
        ('eln', 'ELN\\n(:8002)'),
        ('db', 'PostgreSQL\\n(eln)'),
    ],
    'messages': [
        (0, 1, '서명 버튼 클릭 + 비밀번호 입력', 'normal'),
        (1, 2, 'POST /api/signatures/sign/:noteId {password, comment}', 'normal'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 4, 'POST /internal/verify-password (x-internal-secret)', 'normal'),
        (4, 3, '{verified: true}', 'return'),
        (3, 1, '[실패] 400 SIGNATURE_PASSWORD_INVALID', 'error'),
        (3, 7, 'GET /api/notes/:noteId (내부 호출)', 'normal'),
        (7, 3, '노트 정보 (authorId, teamId, status)', 'return'),
        (3, 3, '권한 검증 (자기서명 차단, reviewer/admin 확인)', 'self'),
        (3, 5, 'SELECT 이전 서명 (prevHash 조회)', 'normal'),
        (3, 3, 'sha256(noteId:signerId:timestamp:prevHash:comment)', 'self'),
        (3, 5, 'INSERT signature (해시체인)', 'normal'),
        (3, 5, "INSERT auditLog (action:'signed')", 'normal'),
        (3, 6, 'XADD labnote:events * type=NOTE_SIGNED', 'normal'),
        (3, 7, '[폴백] PATCH /status {signed} x-user-role:system', 'async'),
        (3, 2, '201 {ok:true, data:signature}', 'return'),
        (2, 1, '응답 전달', 'return'),
        (1, 0, '서명 완료 안내', 'normal'),
        (6, 7, '--- 비동기 이벤트 소비 ---', 'normal'),
        (7, 8, "UPDATE note SET status='signed'", 'normal'),
        (7, 8, "INSERT noteStatusHistory (in_progress→signed)", 'normal'),
        (7, 6, 'XACK labnote:events (처리 완료)', 'normal'),
    ],
    'regions': [
        (18, 21, 'async (Consumer Group: eln-service)', '#F3E5F5'),
    ],
})

# 5. 관리자 잠금 해제
diagrams.append({
    'title': '5. 관리자 잠금 해제',
    'sheet': '5. 잠금 해제',
    'participants': [
        ('actor', '관리자\\n(Admin)'),
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('eln', 'ELN Service\\n(:8002)'),
        ('auth', 'Auth\\n(:8001)'),
        ('db', 'PostgreSQL\\n(eln)'),
        ('signature', 'Sig-Audit\\n(:8003)'),
    ],
    'messages': [
        (0, 1, '잠금 해제 클릭 + 비밀번호/사유 입력', 'normal'),
        (1, 2, 'POST /api/notes/:id/admin-unlock {adminPassword, reason}', 'normal'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 3, 'requireRole(ADMIN) + requirePermission(NOTE_UNLOCK)', 'self'),
        (3, 5, 'SELECT note WHERE id AND orgId', 'normal'),
        (3, 1, '[없음] 404 NOTE_NOT_FOUND', 'error'),
        (3, 1, '[locked 아님] 400 "잠긴 상태가 아닙니다"', 'error'),
        (3, 4, 'POST /internal/verify-password (x-internal-secret)', 'normal'),
        (4, 3, '{verified: true}', 'return'),
        (3, 1, '[실패] 400 NOTE_ADMIN_PASSWORD_INVALID', 'error'),
        (3, 5, "UPDATE note SET status='draft'", 'normal'),
        (3, 5, 'INSERT noteStatusHistory (locked→draft, isAdminAction:true)', 'normal'),
        (3, 6, '[비동기] 감사로그 note.admin_unlocked', 'async'),
        (3, 6, '[비동기] 알림 발송 (작성자≠관리자 시)', 'async'),
        (3, 2, "200 {ok:true, data:note, message:'잠금 해제'}", 'return'),
        (2, 1, '응답 전달', 'return'),
        (1, 0, '상태 변경 반영 (locked → draft)', 'normal'),
    ],
    'regions': [
        (12, 13, 'par (비동기)', '#FFF9C4'),
    ],
})

# 6. PDF/ZIP 내보내기
diagrams.append({
    'title': '6. PDF/ZIP 내보내기',
    'sheet': '6. PDF 내보내기',
    'participants': [
        ('actor', '사용자'),
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('signature', 'Sig-Audit\\n(:8003)'),
        ('bullmq', 'BullMQ\\n(Redis 큐)'),
        ('worker', 'Export\\nWorker'),
        ('eln', 'ELN\\n(:8002)'),
        ('puppeteer', 'Puppeteer'),
        ('file', 'File Service\\n(:8008)'),
        ('minio', 'MinIO'),
        ('redis', 'Redis\\n(pub/sub)'),
    ],
    'messages': [
        (0, 1, 'PDF 내보내기 클릭', 'normal'),
        (1, 2, 'POST /api/export/pdf/:noteId', 'normal'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 4, "exportQueue.add('pdf', {jobId, noteId})", 'normal'),
        (3, 2, "202 Accepted {jobId, status:'pending'}", 'return'),
        (2, 1, '응답 전달', 'return'),
        (1, 2, 'SSE 연결 GET /api/events/exports', 'normal'),
        (2, 10, 'SUBSCRIBE export-status', 'normal'),
        (4, 5, '작업 수신 (concurrency:2)', 'normal'),
        (5, 10, 'PUBLISH export-status {progress:10%}', 'normal'),
        (5, 6, 'GET /api/notes/:noteId (내부 호출)', 'normal'),
        (5, 5, 'Handlebars 템플릿 컴파일', 'self'),
        (5, 7, 'HTML → PDF 변환 (A4)', 'normal'),
        (7, 5, 'PDF Buffer', 'return'),
        (5, 10, 'PUBLISH export-status {progress:80%}', 'normal'),
        (5, 8, 'POST /api/exports/internal/upload', 'normal'),
        (8, 9, "PutObject (bucket:'labnote-exports')", 'normal'),
        (8, 9, 'PresignedGetObject (24시간 유효)', 'normal'),
        (8, 5, '{fileId, downloadUrl}', 'return'),
        (5, 10, "PUBLISH export-status {completed, 100%}", 'normal'),
        (10, 2, 'SSE event → Frontend', 'normal'),
        (1, 0, '다운로드 버튼 활성화', 'normal'),
        (0, 1, '다운로드 클릭', 'normal'),
        (1, 9, 'GET presigned URL (직접 다운로드)', 'normal'),
    ],
})

# 7. 통합 검색
diagrams.append({
    'title': '7. 통합 검색',
    'sheet': '7. 통합 검색',
    'participants': [
        ('actor', '사용자'),
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('search', 'Search Service\\n(:8006)'),
        ('redis', 'Redis\\n(캐시)'),
        ('opensearch', 'OpenSearch'),
        ('db', 'PostgreSQL\\n(search)'),
    ],
    'messages': [
        (0, 1, '검색어 타이핑 ("세포 배양")', 'normal'),
        (1, 2, 'GET /api/search/suggest?q=세포 배양', 'normal'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 5, 'multi_match (title^3, tags^2, phrase_prefix)', 'normal'),
        (5, 3, '상위 7건 매칭 결과', 'return'),
        (3, 1, '[{text, domainType, docId}, ...]', 'return'),
        (1, 0, '자동완성 드롭다운 표시', 'normal'),
        (0, 1, '검색 실행 (Enter)', 'normal'),
        (1, 2, 'GET /api/search?q=세포 배양&domainTypes=NOTE,PROTOCOL', 'normal'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 3, '캐시 키 생성 SHA256(q+types+page+userId+orgId)', 'self'),
        (3, 4, 'GET cache:{hash}', 'normal'),
        (3, 1, '[캐시 히트] 즉시 응답 (3분 TTL)', 'return'),
        (3, 5, 'bool 쿼리 (title^4, tags^3, fuzziness:AUTO)', 'normal'),
        (5, 3, '검색 결과 + 하이라이트 + 집계', 'return'),
        (3, 6, '[비동기] INSERT searchHistory', 'async'),
        (3, 4, '[비동기] SET cache:{hash} EX 180', 'async'),
        (3, 2, '{results[], counts:{NOTE:12}, pagination}', 'return'),
        (2, 1, '응답 전달', 'return'),
        (1, 0, '검색 결과 목록 + 도메인별 탭', 'normal'),
    ],
    'regions': [
        (15, 16, 'par (비동기)', '#FFF9C4'),
    ],
})

# 8-1. 수량 입출고
diagrams.append({
    'title': '8-1. 인벤토리 수량 입출고',
    'sheet': '8-1. 수량 입출고',
    'participants': [
        ('actor', '연구자'),
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('inventory', 'Inventory\\n(:8004)'),
        ('db', 'PostgreSQL\\n(inventory)'),
        ('search', 'Search\\n(:8006)'),
    ],
    'messages': [
        (0, 1, '시약 출고 수량 입력 (Ethanol 500ml)', 'normal'),
        (1, 2, "POST /api/inventory/items/:id/quantity {out, 500}", 'normal'),
        (2, 3, '프록시 전달', 'normal'),
        (3, 4, 'SELECT inventoryItem WHERE id AND orgId', 'normal'),
        (3, 3, '현재 수량 확인 (2000ml) → 계산: after=1500', 'self'),
        (3, 1, '[부족] 400 ITEM_STOCK_INSUFFICIENT', 'error'),
        (3, 3, "자동 상태 전환 판단 (0→depleted, >0→available)", 'self'),
        (3, 4, "[TX] UPDATE item SET quantity=1500, status='available'", 'normal'),
        (3, 4, '[TX] INSERT inventoryHistory (out, 2000→1500, -500)', 'normal'),
        (3, 5, '[비동기] POST /api/search/index 재인덱싱', 'async'),
        (3, 2, '200 {ok:true, data:{item, before:2000, after:1500}}', 'return'),
        (2, 1, '응답 전달', 'return'),
        (1, 0, '수량 변경 완료 표시', 'normal'),
    ],
    'regions': [
        (7, 8, 'Transaction', '#E3F2FD'),
    ],
})

# 8-2. 재고 경고
diagrams.append({
    'title': '8-2. 재고 경고 (대시보드)',
    'sheet': '8-2. 재고 경고',
    'participants': [
        ('actor', '관리자'),
        ('frontend', 'Frontend'),
        ('gateway', 'API Gateway\\n(:8000)'),
        ('redis', 'Redis\\n(캐시)'),
        ('inventory', 'Inventory\\n(:8004)'),
        ('db', 'PostgreSQL\\n(inventory)'),
    ],
    'messages': [
        (0, 1, '대시보드 접근', 'normal'),
        (1, 2, 'GET /api/dashboard/org', 'normal'),
        (2, 3, 'GET cache:dashboard:org:{orgId}', 'normal'),
        (2, 1, '[캐시 히트] 즉시 응답 (5분 TTL)', 'return'),
        (2, 4, '[병렬] GET /api/inventory/alerts/low-stock', 'normal'),
        (2, 4, '[병렬] GET /api/inventory/alerts/expiring?days=30', 'normal'),
        (4, 5, "SELECT items WHERE quantity<=minQuantity", 'normal'),
        (4, 5, "SELECT items WHERE expiryDate<=NOW()+30일", 'normal'),
        (4, 2, '재고 부족 항목 목록', 'return'),
        (4, 2, '만료 임박 항목 목록 (daysLeft 포함)', 'return'),
        (2, 2, '응답 집계 (lowStock/expiring 상위 5건)', 'self'),
        (2, 3, 'SET cache:dashboard:org:{orgId} EX 300', 'normal'),
        (2, 1, '대시보드 데이터', 'return'),
        (1, 0, '재고 부족 + 만료 임박 경고 카드 표시', 'normal'),
    ],
    'regions': [
        (4, 9, 'par (Promise.allSettled 병렬 요청)', '#E3F2FD'),
    ],
})


# ── 엑셀 생성 ──

wb = openpyxl.Workbook()
ws_toc = wb.active
ws_toc.title = '목차'

# 목차
TITLE_FONT_XL = Font(name='맑은 고딕', bold=True, size=16, color='1A237E')
ws_toc.cell(row=1, column=1, value='LabNote ELN - 시퀀스 다이어그램').font = TITLE_FONT_XL
ws_toc.cell(row=2, column=1, value='전자연구노트(ELN) 협업 플랫폼의 핵심 흐름 8가지를 시퀀스 다이어그램으로 정리').font = Font(name='맑은 고딕', size=11, color='546E7A')

toc_header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
toc_header_font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for c, h in enumerate(['번호', '시퀀스명', '설명', '주요 서비스'], 1):
    cell = ws_toc.cell(row=4, column=c, value=h)
    cell.font = toc_header_font
    cell.fill = toc_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin

toc_data = [
    ('1-1', '로그인', '이메일/비밀번호 → JWT(Access+Refresh) 발급', 'Auth, Redis, PostgreSQL'),
    ('1-2', '토큰 갱신', 'Access Token 만료 → 자동 갱신', 'Auth, Redis'),
    ('1-3', '로그아웃', '토큰 블랙리스트 등록 + 쿠키 삭제', 'Auth, Redis'),
    ('1-4', 'JWT 검증', 'Gateway의 8단계 JWT 검증 파이프라인', 'API Gateway, Redis'),
    ('2-1', '노트 생성', '연구노트 생성 + 리비전 + 검색 인덱싱 + 감사로그', 'ELN, Search, Sig-Audit'),
    ('2-2', '노트 수정', '노트 수정 + 잠김/서명 체크 + 리비전 생성', 'ELN, Search, Sig-Audit'),
    ('3', '실시간 협업', 'WebSocket 기반 실시간 편집 (last-write-wins)', 'Collab, Redis Pub/Sub'),
    ('4', '전자서명', '비밀번호 검증 → 해시체인 서명 → Redis Stream 상태 전환', 'Sig-Audit, Auth, ELN'),
    ('5', '잠금 해제', '관리자 비밀번호 검증 후 locked→draft 전환', 'ELN, Auth, Sig-Audit'),
    ('6', 'PDF 내보내기', 'BullMQ 큐 → Puppeteer PDF → MinIO 업로드 → SSE 진행률', 'Sig-Audit, File, MinIO'),
    ('7', '통합 검색', 'OpenSearch 자동완성 + 본검색 + Redis 캐싱', 'Search, OpenSearch, Redis'),
    ('8-1', '수량 입출고', '인벤토리 수량 변경 (트랜잭션)', 'Inventory, Search'),
    ('8-2', '재고 경고', '대시보드 재고 부족/만료 경고 (병렬 집계)', 'API Gateway, Inventory'),
]

normal_font = Font(name='맑은 고딕', size=10)
for i, (num, name, desc, svcs) in enumerate(toc_data, 5):
    row_fill = PatternFill(start_color='F5F5F5' if i % 2 == 0 else 'FFFFFF', end_color='F5F5F5' if i % 2 == 0 else 'FFFFFF', fill_type='solid')
    for c, val in enumerate([num, name, desc, svcs], 1):
        cell = ws_toc.cell(row=i, column=c, value=val)
        cell.font = normal_font
        cell.border = thin
        cell.fill = row_fill
        cell.alignment = Alignment(vertical='center')

ws_toc.column_dimensions['A'].width = 8
ws_toc.column_dimensions['B'].width = 18
ws_toc.column_dimensions['C'].width = 55
ws_toc.column_dimensions['D'].width = 35

# 범례
lr = len(toc_data) + 7
ws_toc.cell(row=lr, column=1, value='화살표 색상 범례').font = Font(name='맑은 고딕', bold=True, size=11, color='1A237E')
legend_items = [
    ('실선 검정', '동기 호출 (요청)', '#37474F'),
    ('점선 회색', '응답 (반환)', '#78909C'),
    ('실선 빨강', '에러 / 예외 흐름', '#E53935'),
    ('점선 주황', '비동기 호출 (fire-and-forget)', '#FF9800'),
    ('자기 루프', '내부 처리 (self-call)', '#37474F'),
]
for j, (style, desc, color) in enumerate(legend_items, lr + 1):
    ws_toc.cell(row=j, column=1, value=style).font = Font(name='맑은 고딕', size=10, color=color[1:])
    ws_toc.cell(row=j, column=2, value=desc).font = normal_font

# 각 다이어그램을 이미지로 생성 + 엑셀에 삽입
tmp_dir = tempfile.mkdtemp()
print(f'임시 디렉토리: {tmp_dir}')

for d in diagrams:
    print(f"  그리는 중: {d['title']}")

    fig = draw_sequence_diagram(
        title=d['title'],
        participants=d['participants'],
        messages=d['messages'],
        fig_width=max(18, len(d['participants']) * 3.2),
        regions=d.get('regions', []),
    )

    img_path = os.path.join(tmp_dir, f"{d['sheet']}.png")
    fig.savefig(img_path, dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
    plt.close(fig)

    ws = wb.create_sheet(title=d['sheet'])
    ws.cell(row=1, column=1, value=d['title']).font = Font(name='맑은 고딕', bold=True, size=14, color='1A237E')

    img = XlImage(img_path)
    # 이미지 크기 조정
    img.width = img.width * 0.75
    img.height = img.height * 0.75
    ws.add_image(img, 'A3')

# 이벤트 흐름 시트
ws_ev = wb.create_sheet(title='이벤트 흐름')
ws_ev.cell(row=1, column=1, value='서비스 간 이벤트 흐름 매핑').font = Font(name='맑은 고딕', bold=True, size=14, color='1A237E')

ev_headers = ['이벤트명', '통신 방식', '발행자 (Publisher)', '구독자 (Subscriber)', '트리거 조건']
for c, h in enumerate(ev_headers, 1):
    cell = ws_ev.cell(row=3, column=c, value=h)
    cell.font = toc_header_font
    cell.fill = toc_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin

event_mapping = [
    ('NOTE_SIGNED', 'Redis Stream (labnote:events)', 'Signature-Audit', 'ELN Service', '전자서명 완료'),
    ('note.created/updated', 'HTTP POST (내부)', 'ELN Service', 'Search Service', '노트 CRUD'),
    ('inventory.indexed', 'HTTP POST (내부)', 'Inventory Service', 'Search Service', '인벤토리 CRUD'),
    ('export-status', 'Redis Pub/Sub', 'Export Worker', 'API Gateway (SSE)', '내보내기 진행률'),
    ('labnote:collab', 'Redis Pub/Sub', 'Collab Service', 'Collab Service (다른 인스턴스)', '실시간 편집'),
    ('NOTE_LOCKED/UNLOCKED', 'HTTP POST (내부)', 'ELN Service', 'Signature-Audit (알림)', '상태 변경'),
]

for i, ev in enumerate(event_mapping, 4):
    row_fill = PatternFill(start_color='F5F5F5' if i % 2 == 0 else 'FFFFFF', end_color='F5F5F5' if i % 2 == 0 else 'FFFFFF', fill_type='solid')
    for c, val in enumerate(ev, 1):
        cell = ws_ev.cell(row=i, column=c, value=val)
        cell.font = normal_font
        cell.border = thin
        cell.fill = row_fill

ws_ev.column_dimensions['A'].width = 25
ws_ev.column_dimensions['B'].width = 30
ws_ev.column_dimensions['C'].width = 25
ws_ev.column_dimensions['D'].width = 30
ws_ev.column_dimensions['E'].width = 18

# 저장
output_path = r'C:\vscode\lab\lab-companion-feature-phase1-backend-infra\docs\LabNote_ELN_시퀀스다이어그램.xlsx'
wb.save(output_path)
print(f'\n완료: {output_path}')

# 임시 파일 정리
import shutil
shutil.rmtree(tmp_dir, ignore_errors=True)
print('임시 파일 정리 완료')
