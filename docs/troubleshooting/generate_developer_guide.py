#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""개발자 트러블슈팅 가이드 Excel 생성"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 스타일 =====
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='맑은 고딕', size=16, bold=True, color='1F4E79')
SUBTITLE_FONT = Font(name='맑은 고딕', size=12, bold=True, color='333333')
SECTION_FONT = Font(name='맑은 고딕', size=11, bold=True, color='1F4E79')
NORMAL_FONT = Font(name='맑은 고딕', size=10)
BOLD_FONT = Font(name='맑은 고딕', size=10, bold=True)
SMALL_FONT = Font(name='맑은 고딕', size=9, color='555555')
CODE_FONT = Font(name='Consolas', size=9)
RED_FONT = Font(name='맑은 고딕', size=10, bold=True, color='C00000')
ORANGE_FONT = Font(name='맑은 고딕', size=10, bold=True, color='ED7D31')

LIGHT_RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_ORANGE = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
LIGHT_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
LIGHT_PURPLE = PatternFill(start_color='E8D5F5', end_color='E8D5F5', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)


def header_row(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[row].height = 28
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def data_cell(ws, row, col, value, align='left', font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or NORMAL_FONT
    cell.alignment = left if align == 'left' else (top_left if align == 'top' else center)
    cell.border = thin_border
    if fill:
        cell.fill = fill
    elif row % 2 == 0:
        cell.fill = LIGHT_GRAY
    return cell


def title(ws, text, row=1, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 40


def subtitle(ws, text, row, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.fill = LIGHT_BLUE
    cell.border = thin_border
    ws.row_dimensions[row].height = 28


# ================================================================
# 시트 1: 서비스별 에러코드
# ================================================================
ws1 = wb.active
ws1.title = '서비스별 에러코드'
ws1.sheet_properties.tabColor = 'C00000'

title(ws1, 'LabNote ELN 개발자 트러블슈팅 가이드 - 에러코드 전체 목록')

ws1.merge_cells('A2:G2')
ws1['A2'].value = '모든 에러코드는 @lab/shared의 ErrorCode에 정의됨. 에러 발생 시 응답의 code 필드로 원인을 특정합니다.'
ws1['A2'].font = SMALL_FONT
ws1['A2'].alignment = left

hdrs = ['서비스', '에러 코드 (code)', 'HTTP 상태', '에러 메시지 (사용자에게 표시)', '발생 원인', '해결 방법', '관련 파일']
widths = [14, 28, 10, 30, 30, 35, 28]
header_row(ws1, 4, hdrs, widths)

errors = [
    # ── 공통 ──
    ['공통', 'UNAUTHORIZED', 401, '인증이 필요합니다', 'JWT 토큰 누락 또는 만료', 'requireAuth 미들웨어 확인\n토큰 갱신 로직 확인', 'shared/src/middleware.ts'],
    ['공통', 'FORBIDDEN', 403, '권한이 없습니다', '권한 부족 (Permission 미보유)', 'requirePermission() 설정 확인\n사용자 역할의 권한 목록 확인', 'shared/src/middleware.ts'],
    ['공통', 'NOT_FOUND', 404, '리소스를 찾을 수 없습니다', '존재하지 않는 ID 또는 orgId 불일치', 'orgId 필터 확인 (withOrgScope)\nID 유효성 확인', 'shared/src/errors.ts'],
    ['공통', 'CONFLICT', 409, '리소스 충돌', '유니크 제약 위반 또는 상태 충돌', '중복 데이터 확인\n상태 전환 규칙 확인', 'shared/src/errors.ts'],
    ['공통', 'VALIDATION_ERROR', 400, '입력값이 올바르지 않습니다', 'Zod 스키마 검증 실패', 'DTO 스키마 확인\n요청 body/query/params 형식 확인', 'shared/src/validate.ts'],
    ['공통', 'INTERNAL_ERROR', 500, '서버 내부 오류', '예상치 못한 예외 발생', '서비스 로그 확인\nstack trace 분석', 'shared/src/errors.ts'],
    ['공통', 'SERVICE_UNAVAILABLE', 503, '서비스를 사용할 수 없습니다', '외부 의존성(DB, Redis 등) 연결 실패', 'docker ps로 인프라 상태 확인\nDATABASE_URL, REDIS_URL 확인', 'shared/src/errors.ts'],
    ['공통', 'INTERNAL_AUTH_FAILED', 401, '내부 요청 인증 실패', 'x-internal-secret 헤더 불일치', 'INTERNAL_SECRET 환경변수 일치 여부 확인\n서비스 간 동일 값 사용 중인지 확인', 'shared/src/middleware.ts'],

    # ── Auth ──
    ['auth-service', 'AUTH_INVALID_CREDENTIALS', 401, '이메일 또는 비밀번호가 올바르지 않습니다', '로그인 시 이메일/비밀번호 불일치', 'DB에서 사용자 조회 확인\nbcrypt 해시 비교 로직 확인', 'auth-service/src/controllers/\nauth.controller.ts'],
    ['auth-service', 'AUTH_INACTIVE_ACCOUNT', 403, '비활성 계정입니다', '사용자 isActive=false', 'DB에서 사용자 상태 확인\n관리자가 활성화 필요', 'auth.controller.ts'],
    ['auth-service', 'AUTH_TOKEN_REQUIRED', 401, '토큰이 필요합니다', 'Refresh 토큰 미전달', '요청 body에 refreshToken 포함 여부 확인', 'auth.controller.ts'],
    ['auth-service', 'AUTH_TOKEN_EXPIRED', 401, '토큰이 만료되었습니다', 'Refresh 토큰 유효기간(8시간) 초과', '재로그인 필요\nRefresh 토큰 TTL 설정 확인', 'auth.controller.ts'],
    ['auth-service', 'AUTH_TOKEN_INVALID', 401, '유효하지 않은 토큰', '변조된 토큰 또는 권한 변경 후 기존 토큰 사용', 'JWT_SECRET 일치 여부 확인\nRedis 블랙리스트 확인', 'auth.controller.ts\ntoken-invalidation.ts'],
    ['auth-service', 'AUTH_EMAIL_EXISTS', 409, '이미 등록된 이메일입니다', '회원가입 시 이메일 중복', 'DB unique 제약 정상 동작\n사용자 안내', 'auth.controller.ts'],
    ['auth-service', 'AUTH_NO_ORGANIZATION', 403, '조직 정보가 없습니다', '시스템에 조직이 하나도 없는 상태', 'DB에 Organization 레코드 존재 확인\n시드 데이터 실행 확인', 'auth.controller.ts'],
    ['auth-service', 'AUTH_SELF_DELETE', 400, '자기 자신을 삭제할 수 없습니다', '관리자가 본인 계정 삭제 시도', '정상 동작 — 다른 Admin이 삭제해야 함', 'user.controller.ts'],
    ['auth-service', 'AUTH_ROLE_HAS_USERS', 409, '사용 중인 역할은 삭제할 수 없습니다', '해당 역할에 배정된 사용자 존재', '사용자 역할 변경 후 삭제\n또는 역할 유지', 'role.controller.ts'],
    ['auth-service', 'AUTH_ORG_HAS_USERS', 409, '소속 사용자가 있는 조직은 삭제 불가', '조직에 소속된 사용자 존재', '모든 사용자 이동/삭제 후 조직 삭제', 'org.controller.ts'],
    ['auth-service', 'AUTH_INVALID_ROLE_NAME', 400, '허용되지 않은 역할 이름', 'admin/reviewer/researcher/viewer 외 이름', '허용 역할명 목록 확인', 'role.controller.ts'],
    ['auth-service', 'AUTH_SLUG_EXISTS', 409, '이미 사용 중인 slug', '조직 생성 시 slug 중복', '다른 slug 사용 안내', 'org.controller.ts'],

    # ── ELN ──
    ['eln-service', 'NOTE_NOT_FOUND', 404, '노트를 찾을 수 없습니다', '존재하지 않는 noteId 또는 타 조직 노트', 'orgId 필터 확인\ndeletedAt IS NULL 조건 확인 (소프트삭제)', 'eln-service/src/controllers/\nnote.controller.ts'],
    ['eln-service', 'NOTE_LOCKED', 409, '잠금된 노트는 수정할 수 없습니다', 'locked 상태 노트 수정/삭제 시도', '정상 동작 — Admin 잠금해제 필요', 'note.controller.ts'],
    ['eln-service', 'NOTE_SIGNED', 409, '서명된 노트는 변경할 수 없습니다', 'signed 상태 노트 수정/삭제 시도', '정상 동작 — signed는 불변', 'note.controller.ts'],
    ['eln-service', 'NOTE_STATUS_TRANSITION', 400, '허용되지 않은 상태 전환입니다', 'ALLOWED_STATUS_TRANSITIONS에 없는 전환', 'dtos/note.dto.ts의 전환 맵 확인\nDB 트리거 check_note_status_transition도 확인', 'note.dto.ts\nnote.controller.ts'],
    ['eln-service', 'NOTE_PERMISSION_DENIED', 403, '노트 접근 권한이 없습니다', '타인 노트 수정/삭제 시도', '소유자/Admin 검증 로직 확인\nrequireOwnerOrAdmin 미들웨어', 'note.controller.ts'],
    ['eln-service', 'NOTE_LOCK_ROLE_REQUIRED', 403, '잠금은 Reviewer 이상만 가능합니다', 'Researcher가 in_progress→locked 시도', '정상 동작 — x-user-role 검증', 'note.controller.ts'],
    ['eln-service', 'NOTE_ADMIN_PASSWORD_INVALID', 401, '비밀번호가 올바르지 않습니다', '관리자 잠금해제 시 비밀번호 불일치', 'auth-service의 internal/verify-password 연동 확인', 'note.controller.ts'],
    ['eln-service', 'TEMPLATE_NOT_FOUND', 404, '템플릿을 찾을 수 없습니다', '존재하지 않는 templateId', 'orgId 필터 확인', 'note.controller.ts'],

    # ── Signature ──
    ['signature-audit', 'SIGNATURE_PASSWORD_INVALID', 401, '비밀번호가 올바르지 않습니다', '서명 시 비밀번호 검증 실패', 'auth-service verify-password API 연동 확인', 'signature-audit-service/src/\ncontrollers/signature.controller.ts'],
    ['signature-audit', 'SIGNATURE_ALREADY_REVOKED', 409, '이미 취소된 서명입니다', '취소된 서명 재취소 시도', 'signature.status 확인', 'signature.controller.ts'],
    ['signature-audit', 'SIGNATURE_ADMIN_ONLY', 403, 'Admin만 가능한 작업입니다', '비Admin의 서명 취소 시도', 'requireRole(ADMIN) 미들웨어 확인', 'signature.controller.ts'],

    # ── Inventory ──
    ['inventory-service', 'ITEM_NOT_FOUND', 404, '항목을 찾을 수 없습니다', '존재하지 않는 itemId 또는 타 조직', 'orgId 필터 확인', 'inventory-service/src/controllers/\ninventory.controller.ts'],
    ['inventory-service', 'ITEM_BARCODE_EXISTS', 409, '이미 등록된 바코드입니다', '바코드 유니크 제약 위반', 'DB unique index 확인', 'inventory.controller.ts'],
    ['inventory-service', 'ITEM_STOCK_INSUFFICIENT', 400, '재고가 부족합니다', '출고 수량 > 현재 재고', '정상 동작 — 음수 재고 방지', 'inventory.controller.ts'],
    ['inventory-service', 'CATEGORY_EXISTS', 409, '이미 존재하는 카테고리입니다', '카테고리명 중복', 'orgId+name 유니크 제약 확인', 'inventory.controller.ts'],

    # ── File ──
    ['file-service', 'FILE_NOT_FOUND', 404, '파일을 찾을 수 없습니다', '존재하지 않는 fileId 또는 타 조직', 'orgId 필터 확인\ndeletedAt 소프트삭제 확인', 'file-service/src/controllers/\nfile.controller.ts'],
    ['file-service', 'FILE_BLOCKED_MIME', 400, '허용되지 않는 파일 형식입니다', '.exe, .sh, .bat 등 실행파일 업로드', '차단 MIME 목록 확인', 'file.controller.ts'],
    ['file-service', 'FILE_UPLOAD_FAILED', 502, '파일 업로드에 실패했습니다', 'MinIO 연결 실패 또는 스토리지 오류', 'MinIO 상태 확인\nMINIO_ENDPOINT, MINIO_PORT 확인', 'file.controller.ts'],
    ['file-service', 'FILE_METADATA_FAILED', 500, '파일 정보 저장에 실패했습니다', 'MinIO 업로드 성공 후 DB 저장 실패', 'DB 연결 확인\n자동 롤백(MinIO 삭제) 로그 확인', 'file.controller.ts'],

    # ── Export ──
    ['signature-audit', 'EXPORT_NOT_FOUND', 404, '내보내기 작업을 찾을 수 없습니다', '존재하지 않는 jobId', 'DB 확인', 'export.controller.ts'],
    ['signature-audit', 'EXPORT_NOT_COMPLETED', 400, '아직 완료되지 않은 작업입니다', 'PENDING/PROCESSING 상태에서 다운로드 시도', '정상 동작 — 완료 대기 필요', 'export.controller.ts'],
    ['signature-audit', 'EXPORT_NOT_CANCELLABLE', 409, '취소할 수 없는 작업입니다', 'PROCESSING 이후 상태에서 취소 시도', '정상 동작 — 처리 중 작업은 취소 불가', 'export.controller.ts'],

    # ── Search ──
    ['search-service', 'SEARCH_FAVORITE_EXISTS', 409, '이미 즐겨찾기에 추가되었습니다', '동일 docType+docId 중복 즐겨찾기', '정상 동작 — userId+docType+docId 유니크', 'search-service/src/controllers/\nfavorite.controller.ts'],
    ['search-service', 'SEARCH_PERMISSION_DENIED', 403, '검색 이력 접근 권한이 없습니다', '타인의 검색 이력/즐겨찾기 삭제 시도', 'userId 소유권 검증 로직 확인', 'history.controller.ts'],

    # ── Notification ──
    ['signature-audit', 'NOTIFICATION_NOT_FOUND', 404, '알림을 찾을 수 없습니다', '존재하지 않는 알림 ID', 'orgId+userId 필터 확인', 'notification.controller.ts'],
]

for r, data in enumerate(errors, 5):
    for c, val in enumerate(data, 1):
        align = 'top' if c in [4, 5, 6, 7] else 'center'
        font = CODE_FONT if c in [2, 7] else (BOLD_FONT if c == 1 else None)
        # 서비스별 색상
        svc_fill = None
        if c == 1:
            svc = data[0]
            if svc == '공통':
                svc_fill = LIGHT_BLUE
            elif 'auth' in svc:
                svc_fill = LIGHT_GREEN
            elif 'eln' in svc:
                svc_fill = LIGHT_ORANGE
            elif 'signature' in svc:
                svc_fill = LIGHT_PURPLE
        data_cell(ws1, r, c, val, align, font, svc_fill)
    ws1.row_dimensions[r].height = 45


# ================================================================
# 시트 2: 서비스 간 통신 장애
# ================================================================
ws2 = wb.create_sheet('서비스 간 통신 장애')
ws2.sheet_properties.tabColor = 'ED7D31'

title(ws2, '서비스 간 통신 장애 트러블슈팅')

hdrs = ['No', '장애 시나리오', '영향', '증상 / 에러 로그', '원인', '해결 방법', '예방 조치']
widths = [5, 25, 18, 35, 25, 40, 25]
header_row(ws2, 3, hdrs, widths)

comm_issues = [
    [1, 'Redis Stream\nNOTE_SIGNED 이벤트 미소비',
     '서명 완료 후 노트 상태가\nsigned로 안 바뀜',
     '[eln-service] 로그:\n"event consumer error"\n또는 이벤트 로그 자체가 없음',
     '1. Redis 연결 끊김\n2. Consumer group 미생성\n3. eln-service 재시작 중 이벤트 유실',
     '1. Redis 상태 확인: docker exec redis redis-cli ping\n2. Consumer group 확인:\n   docker exec redis redis-cli XINFO GROUPS labnote:events\n3. eln-service 재시작 (미처리 이벤트 자동 복구 — XCLAIM)\n4. 최대 5회 재시도 후 dead letter 처리',
     'Redis 모니터링 설정\neln-service 헬스체크에\nevent consumer 상태 포함됨'],

    [2, 'Redis Stream → HTTP 폴백\n둘 다 실패',
     '서명 생성은 되었으나\n노트 상태 미반영 (불일치)',
     '[signature-audit] 로그:\n"publish failed" +\n"HTTP fallback failed"',
     '1. Redis + eln-service 동시 장애\n2. INTERNAL_SECRET 불일치\n3. eln-service URL 잘못 설정',
     '1. Redis 상태 확인 및 재시작\n2. eln-service 상태 확인 및 재시작\n3. INTERNAL_SECRET 환경변수 일치 확인\n4. ELN_SERVICE_URL 환경변수 확인\n5. 수동 상태 수정: DB에서 직접 UPDATE (최후 수단)',
     'Redis + eln-service 동시 장애는\n극히 드문 케이스\n정기 헬스체크로 예방'],

    [3, 'x-internal-secret 불일치',
     '서비스 간 내부 API 호출\n전체 실패',
     '[대상 서비스] 로그:\n"내부 요청 인증 실패"\nHTTP 401 INTERNAL_AUTH_FAILED',
     '서비스별 INTERNAL_SECRET\n환경변수 값이 서로 다름',
     '1. docker compose config | grep INTERNAL_SECRET\n   로 모든 서비스의 값 확인\n2. .env 파일에서 통일\n3. 변경 후 전체 서비스 재시작:\n   docker compose up -d --build',
     '.env 파일에 단일 변수로 관리\ndocker-compose.yml에서\n${INTERNAL_SECRET} 참조'],

    [4, 'JWT_SECRET 불일치',
     '로그인은 되지만\n다른 서비스에서 인증 실패',
     '[api-gateway] 또는 [collab-service]:\n"JWT verification failed"\nHTTP 401',
     'auth-service(발급)와\napi-gateway/collab-service(검증)의\nJWT_SECRET 값이 다름',
     '1. docker compose config | grep JWT_SECRET\n2. auth, api-gateway, collab 세 서비스의\n   JWT_SECRET 동일 여부 확인\n3. .env에서 통일 후 세 서비스 재시작',
     'JWT_SECRET은 반드시\n모든 관련 서비스에서 동일값 사용'],

    [5, 'auth-service\nverify-password API 실패',
     '전자서명, 관리자 잠금해제\n비밀번호 검증 불가',
     '[signature-audit] 또는 [eln-service]:\n"auth service verify failed"\nHTTP 502',
     '1. auth-service 중단\n2. AUTH_SERVICE_URL 잘못 설정\n3. 네트워크 분리',
     '1. auth-service 상태 확인\n2. AUTH_SERVICE_URL 확인:\n   docker compose config | grep AUTH_SERVICE_URL\n3. Docker 내부 DNS 확인:\n   docker exec eln-service ping auth-service',
     'auth-service는 다른 서비스의\n인증 핵심 — 최우선 복구 대상'],

    [6, 'search-service 인덱싱 실패\n(노트/인벤토리 생성 후)',
     '새로 생성한 데이터가\n검색 결과에 안 나옴',
     '[eln-service] 또는 [inventory-service]:\n"search index failed"\n(경고 수준, 서비스는 정상)',
     '1. search-service 중단\n2. OpenSearch 중단\n3. SEARCH_SERVICE_URL 오류',
     '1. search-service 상태 확인\n2. OpenSearch 상태: curl http://localhost:9200/_cluster/health\n3. search-service 재시작 (인덱스 자동 재생성)\n※ 인덱싱 실패는 CRUD에 영향 없음 (비동기)',
     '검색 인덱싱은 비동기 처리\nCRUD와 분리되어 있어\n검색 장애가 핵심 기능을 막지 않음'],

    [7, 'BullMQ 작업 큐 정체\n(Export 작업)',
     'PDF/ZIP 내보내기가\n"변환 중..." 무한 대기',
     '[signature-audit] worker 로그:\n작업 처리 로그 없음\n또는 "job failed" 반복',
     '1. Redis 연결 실패\n2. Worker 프로세스 크래시\n3. Puppeteer(PDF 변환) 오류',
     '1. Redis 상태 확인\n2. signature-audit-service 로그 확인\n3. 서비스 재시작 (미완료 작업 자동 재처리)\n4. BullMQ 큐 상태:\n   docker exec redis redis-cli LLEN bull:labnote-export:wait',
     'BullMQ: 3회 재시도 (5s 지수 백오프)\n재시작 시 PENDING 작업 자동 복구'],

    [8, 'WebSocket 연결 실패\n(collab-service)',
     '실시간 협업 편집 불가\n단독 편집 모드로 전환',
     '브라우저 콘솔:\n"WebSocket connection failed"\n또는 101 Upgrade 실패',
     '1. collab-service 중단\n2. 리버스 프록시가 WS 차단\n3. JWT_SECRET 불일치',
     '1. collab-service 상태 확인 (포트 8009)\n2. 리버스 프록시(Nginx 등) WebSocket 설정:\n   proxy_set_header Upgrade $http_upgrade;\n   proxy_set_header Connection "upgrade";\n3. JWT_SECRET 일치 확인',
     'Nginx 사용 시 WebSocket\nupgrade 설정 필수'],

    [9, 'MinIO 연결 실패',
     '파일 업/다운로드 불가\nPresigned URL 발급 실패',
     '[file-service] 로그:\n"MinIO upload failed"\nHTTP 502 FILE_UPLOAD_FAILED',
     '1. MinIO 컨테이너 중단\n2. MINIO_ENDPOINT 오류\n3. 인증정보(ACCESS_KEY) 불일치\n4. 디스크 용량 부족',
     '1. MinIO 상태: docker ps | grep minio\n2. 콘솔 접속: http://서버IP:9001\n3. MINIO_ACCESS_KEY, MINIO_SECRET_KEY 확인\n4. 디스크 용량 확인',
     'MinIO 헬스: /minio/health/live\n디스크 모니터링 설정 권장'],

    [10, 'API Gateway → 백엔드\n프록시 실패',
     '특정 API만 502 에러\n또는 전체 API 502',
     '[api-gateway] 로그:\n"upstream service error"\n또는 "ECONNREFUSED"',
     '1. 대상 서비스 중단\n2. 서비스 URL 환경변수 오류\n3. Docker 내부 네트워크 분리',
     '1. 대상 서비스 상태 확인: docker ps\n2. 환경변수 확인:\n   docker compose config | grep SERVICE_URL\n3. Docker 네트워크:\n   docker network inspect services_default\n4. 해당 서비스 재시작',
     'Gateway가 프록시하는 모든 URL은\ndocker-compose.yml의\n환경변수로 관리'],
]

for r, data in enumerate(comm_issues, 4):
    for c, val in enumerate(data, 1):
        data_cell(ws2, r, c, val, 'top')
    ws2.row_dimensions[r].height = 100


# ================================================================
# 시트 3: DB 마이그레이션 이슈
# ================================================================
ws3 = wb.create_sheet('DB 마이그레이션 이슈')
ws3.sheet_properties.tabColor = '548235'

title(ws3, 'DB / Prisma 마이그레이션 트러블슈팅')

hdrs = ['No', '문제', '에러 메시지 / 증상', '원인', '해결 방법', '주의사항', '관련 서비스']
widths = [5, 22, 35, 25, 45, 25, 14]
header_row(ws3, 3, hdrs, widths)

db_issues = [
    [1, 'migrate deploy 실패\n"migration not found"',
     'Error: P3009\nMigration directory not found',
     'prisma/migrations/ 디렉토리가\n컨테이너에 복사되지 않음',
     '1. Dockerfile에서 prisma/ 디렉토리 COPY 확인\n2. .dockerignore에 prisma/ 포함 여부 확인\n3. docker compose up -d --build <서비스> 로 재빌드',
     '마이그레이션 파일은 코드와 함께\n반드시 이미지에 포함되어야 함',
     '모든 서비스'],

    [2, 'DB 스키마 불일치\n"column does not exist"',
     'PrismaClientKnownRequestError\nP2022: Column not found',
     '코드 변경 후 마이그레이션 미실행',
     '1. docker exec <서비스> npx prisma migrate deploy\n2. 개발 중이라면: npx prisma migrate dev --name <설명>\n3. 마이그레이션 실패 시 수동 SQL 확인',
     'migrate dev는 개발용\nprodution에서는 migrate deploy 사용\n마이그레이션 파일 직접 수정 금지',
     '해당 서비스'],

    [3, 'DB 트리거 충돌\ncheck_note_status_transition',
     'PostgreSQL ERROR:\nInvalid status transition\n(DB 레벨 차단)',
     'DB 트리거가 잘못된 상태 전환을 차단\n(정상 안전장치 작동)',
     '1. 전환 맵 확인:\n   ALLOWED_STATUS_TRANSITIONS (일반)\n   SYSTEM_STATUS_TRANSITIONS (서명 서비스)\n2. system 역할은 in_progress→signed 전환 가능\n3. 트리거 수정 절대 금지 — 코드에서 해결',
     'DB 트리거는 최종 안전장치\n절대 수정/삭제 금지\n코드의 전환 맵이 우선 적용',
     'eln-service'],

    [4, 'Connection pool exhausted\n커넥션 풀 고갈',
     'PrismaClientInitializationError\nCan\'t reach database server\nToo many connections',
     '1. connection_limit 초과\n2. 커넥션 미반환 (leak)\n3. 서비스 다중 인스턴스',
     '1. DATABASE_URL의 connection_limit 확인 (기본 3)\n2. pool_timeout 확인 (기본 10초)\n3. DB 커넥션 수:\n   docker exec postgres psql -U labnote\n   -c "SELECT count(*) FROM pg_stat_activity"\n4. 서비스 재시작으로 커넥션 초기화',
     'PostgreSQL max_connections=30\n서비스당 connection_limit=3\n(9서비스 × 3 = 27)',
     '모든 서비스'],

    [5, 'Prisma 클라이언트\n버전 불일치',
     '@prisma/client version mismatch\nprisma와 @prisma/client 버전 상이',
     'package.json의 prisma와\n@prisma/client 버전이 다름',
     '1. npx prisma generate 실행\n2. 또는 @prisma/client 버전 맞추기\n3. docker compose up -d --build <서비스>',
     '두 패키지는 항상 동일 버전이어야 함',
     '해당 서비스'],

    [6, 'Soft delete 데이터\n조회 이상',
     '삭제한 노트가 계속 조회됨\n또는 삭제된 노트가 안 보여야 하는데 보임',
     'eln-service의 soft delete 미들웨어가\ndeletedAt IS NULL 자동 필터링\n다른 서비스에는 미적용',
     '1. eln-service: 자동 필터링 동작 확인\n   (prisma.ts의 $use 미들웨어)\n2. 다른 서비스: where 절에 deletedAt: null 명시 필요\n3. 삭제된 데이터 포함 조회 시:\n   where: { deletedAt: { not: null } } 명시',
     'eln-service만 soft delete 미들웨어 적용\n다른 서비스는 수동 필터링 필요',
     'eln-service\n(다른 서비스 주의)'],

    [7, 'UUID 충돌\n또는 시퀀스 오류',
     'Unique constraint violation\nP2002',
     '1. UUID 생성기 충돌 (극히 드뭄)\n2. 수동 INSERT 시 ID 중복\n3. 시드 데이터 재실행',
     '1. 시드 데이터: upsert 사용 확인\n2. 수동 INSERT 시 gen_random_uuid() 사용\n3. 시퀀스 리셋 필요 시:\n   SELECT setval(pg_get_serial_sequence(...), max(id))',
     'Prisma의 기본 UUID 생성은\ncuid/uuid를 사용하므로 충돌 극히 드뭄',
     '모든 서비스'],
]

for r, data in enumerate(db_issues, 4):
    for c, val in enumerate(data, 1):
        data_cell(ws3, r, c, val, 'top', CODE_FONT if c == 3 else None)
    ws3.row_dimensions[r].height = 100


# ================================================================
# 시트 4: 성능 이슈
# ================================================================
ws4 = wb.create_sheet('성능 이슈')
ws4.sheet_properties.tabColor = '4472C4'

title(ws4, '성능 이슈 트러블슈팅')

hdrs = ['No', '증상', '영향 범위', '진단 방법', '원인', '해결 방법', '모니터링']
widths = [5, 22, 14, 35, 25, 40, 22]
header_row(ws4, 3, hdrs, widths)

perf_issues = [
    [1, 'API 응답 전반적 느림\n(1초 이상)',
     '전체 사용자',
     '1. Jaeger UI (http://서버IP:16686)에서\n   느린 요청 트레이스 확인\n2. docker stats 로 리소스 확인\n3. PostgreSQL slow query:\n   docker exec postgres psql -U labnote\n   -c "SELECT * FROM pg_stat_activity\n   WHERE state=\'active\'"',
     '1. DB 쿼리 느림 (인덱스 누락)\n2. 서버 CPU/메모리 부족\n3. 커넥션 풀 고갈',
     '1. Jaeger에서 병목 서비스 특정\n2. DB 인덱스 추가 (orgId 복합 인덱스)\n3. 서버 리소스 증설\n4. connection_limit 조정',
     'Jaeger 트레이싱\ndocker stats\npg_stat_activity'],

    [2, 'OpenSearch 검색 느림\n(3초 이상)',
     '검색 사용자',
     '1. curl http://localhost:9200/_cluster/health\n2. curl http://localhost:9200/_cat/indices\n   로 인덱스 크기 확인\n3. search-service 로그에서\n   쿼리 소요 시간 확인',
     '1. 인덱스 크기 과대\n2. OpenSearch JVM 메모리 부족\n3. 복잡한 한국어 분석 쿼리',
     '1. OPENSEARCH_JAVA_OPTS 메모리 증가\n   (기본 256MB → 512MB)\n2. 인덱스 최적화:\n   curl -X POST localhost:9200/lab_search/_forcemerge\n3. Redis 검색 캐시 활용 확인',
     'OpenSearch /_cluster/health\n인덱스 크기 모니터링'],

    [3, 'Redis 메모리 초과\n응답 지연/실패',
     '캐시/세션 관련 전체',
     '1. docker exec redis redis-cli info memory\n2. used_memory_human 확인\n3. maxmemory=128MB 대비 사용량',
     '1. 캐시 데이터 과다 축적\n2. 이벤트 스트림 크기 초과\n3. 세션 데이터 누적',
     '1. 캐시 정리:\n   docker exec redis redis-cli FLUSHDB\n   (세션/캐시 초기화 — 재로그인 필요)\n2. maxmemory 증가 (docker-compose.yml)\n3. allkeys-lru 정책이 자동 eviction 수행',
     'Redis info memory\nused_memory_human\nevicted_keys 카운터'],

    [4, 'PDF 내보내기 느림\n(대용량 노트)',
     '내보내기 사용자',
     '1. signature-audit-service 로그에서\n   worker 처리 시간 확인\n2. BullMQ 대기 큐 길이:\n   docker exec redis redis-cli\n   LLEN bull:labnote-export:wait',
     '1. Puppeteer PDF 변환 메모리 부족\n2. 대용량 노트 (이미지 다수)\n3. 동시 내보내기 작업 과다',
     '1. 서버 메모리 확인 (Puppeteer는 메모리 다소 사용)\n2. 동시 작업 수 제한 확인\n3. BullMQ worker concurrency 조정\n4. 대용량 노트는 ZIP으로 분할 안내',
     'BullMQ 대기 큐 길이\nworker 처리 시간 로그'],

    [5, 'WebSocket 연결 수\n과다로 메모리 증가',
     '협업 편집 사용자',
     '1. http://서버IP:8009/health\n   에서 rooms, clients 수 확인\n2. docker stats collab-service\n   메모리 사용량 확인',
     '1. 많은 사용자 동시 편집\n2. 브라우저 탭 다수 열어둠\n3. WebSocket 연결 미정리',
     '1. collab-service 재시작 (빈 룸 자동 정리됨)\n2. 불필요한 탭 닫기 안내\n3. 필요 시 collab-service 메모리 한도 증가',
     'collab-service /health\nrooms 카운터'],

    [6, 'Rate Limit 429 에러\n빈발',
     '특정 사용자 또는 IP',
     '1. api-gateway 로그에서\n   429 응답 확인\n2. 해당 사용자/IP 특정',
     '1. 자동화 스크립트의 과도한 API 호출\n2. Rate Limit 설정이 너무 낮음\n3. 클라이언트 재시도 로직 과다',
     '1. 글로벌: 200 req/min (기본)\n2. 인증: 20 req/min\n3. 내보내기: 10 req/min\n4. 필요 시 api-gateway/src/index.ts에서\n   Rate Limit 값 조정 후 재빌드',
     'api-gateway 로그\n429 응답 빈도 모니터링'],

    [7, 'PostgreSQL\n디스크 용량 증가',
     '전체 시스템',
     '1. docker exec postgres df -h\n2. docker system df\n3. DB 크기:\n   docker exec postgres psql -U labnote\n   -c "SELECT pg_size_pretty(pg_database_size(\'labnote\'))"',
     '1. NoteRevision 누적 (매 저장마다 리비전)\n2. AuditLog 누적 (불변 기록)\n3. VACUUM 미실행',
     '1. VACUUM ANALYZE 실행:\n   docker exec postgres psql -U labnote\n   -c "VACUUM ANALYZE"\n2. 오래된 리비전 아카이브 검토\n3. 디스크 증설 계획',
     'DB 크기 모니터링\nVACUUM 주기적 실행\n(주 1회 권장)'],
]

for r, data in enumerate(perf_issues, 4):
    for c, val in enumerate(data, 1):
        data_cell(ws4, r, c, val, 'top', CODE_FONT if c == 4 else None)
    ws4.row_dimensions[r].height = 110


# ================================================================
# 시트 5: 환경변수 참조표
# ================================================================
ws5 = wb.create_sheet('환경변수 참조표')
ws5.sheet_properties.tabColor = '7030A0'

title(ws5, '환경변수 전체 참조표')

ws5.merge_cells('A2:G2')
ws5['A2'].value = '모든 환경변수는 services/.env 파일 또는 docker-compose.yml에서 관리됩니다. 서비스 간 동일해야 하는 값에 특히 주의하세요.'
ws5['A2'].font = SMALL_FONT
ws5['A2'].alignment = left

hdrs = ['구분', '변수명', '사용 서비스', '설명', '기본값', '동일값 필수', '비고']
widths = [12, 22, 22, 25, 20, 10, 20]
header_row(ws5, 4, hdrs, widths)

envs = [
    ['인증 (핵심)', 'JWT_SECRET', 'auth, api-gateway,\ncollab-service', 'JWT 토큰 서명/검증 키', '(필수 설정)', 'O (필수)',
     '불일치 시 인증 전체 실패'],
    ['인증 (핵심)', 'INTERNAL_SECRET', '모든 백엔드 서비스', '서비스 간 내부 인증 헤더', 'dev-internal-secret', 'O (필수)',
     '불일치 시 서비스 간\n통신 전체 실패'],
    ['DB', 'POSTGRES_USER', 'postgres, 모든 서비스', 'DB 사용자명', 'labnote', 'O',
     'DATABASE_URL에 포함'],
    ['DB', 'POSTGRES_PASSWORD', 'postgres, 모든 서비스', 'DB 비밀번호', '(필수 설정)', 'O',
     'DATABASE_URL에 포함'],
    ['DB', 'DATABASE_URL', '모든 백엔드 서비스\n(scheduler 제외)', 'Prisma DB 연결 문자열', 'postgresql://user:pass\n@postgres:5432/labnote\n?schema=<서비스별>', '-',
     '서비스별 schema= 값이 다름'],
    ['Redis', 'REDIS_URL', 'auth, eln, search,\ninventory, signature,\ncollab, api-gateway, file', 'Redis 연결 URL', 'redis://redis:6379', 'O',
     'Redis 없이도 일부 기능 동작\n(graceful degradation)'],
    ['스토리지', 'MINIO_ENDPOINT', 'file-service', 'MinIO 호스트', 'minio', '-', 'Docker 내부 DNS'],
    ['스토리지', 'MINIO_PORT', 'file-service', 'MinIO 포트', '9000', '-', ''],
    ['스토리지', 'MINIO_ACCESS_KEY', 'file-service, minio', 'MinIO 인증 키', '(필수 설정)', 'O',
     'MINIO_ROOT_USER와 동일'],
    ['스토리지', 'MINIO_SECRET_KEY', 'file-service, minio', 'MinIO 비밀 키', '(필수 설정)', 'O',
     'MINIO_ROOT_PASSWORD와 동일'],
    ['스토리지', 'MINIO_BUCKET', 'file-service', '파일 저장 버킷명', 'labnote-files', '-', '자동 생성'],
    ['스토리지', 'MINIO_EXPORTS_BUCKET', 'file-service', '내보내기 저장 버킷명', 'labnote-exports', '-', '7일 자동 만료 설정'],
    ['검색', 'OPENSEARCH_URL', 'search-service', 'OpenSearch 연결 URL', 'http://opensearch:9200', '-', ''],
    ['서비스 URL', 'AUTH_SERVICE_URL', 'api-gateway, eln,\nsignature, file', 'auth-service 내부 URL', 'http://auth-service:8001', '-', 'Docker 내부 DNS'],
    ['서비스 URL', 'ELN_SERVICE_URL', 'api-gateway, signature,\nfile', 'eln-service 내부 URL', 'http://eln-service:8002', '-', ''],
    ['서비스 URL', 'SIGNATURE_SERVICE_URL', 'api-gateway, eln', 'signature 내부 URL', 'http://signature-audit\n-service:8003', '-', ''],
    ['서비스 URL', 'FILE_SERVICE_URL', 'api-gateway, signature', 'file-service 내부 URL', 'http://file-service:8008', '-', ''],
    ['CORS', 'CORS_ORIGIN', 'api-gateway', '허용 Origin (프론트엔드)', 'http://localhost:5173', '-', '쉼표로 복수 지정 가능'],
    ['포트', 'GATEWAY_PORT', 'api-gateway', 'Gateway 외부 포트', '8000', '-', '방화벽 개방 필요'],
    ['초기 비밀번호', 'ADMIN_INITIAL_PASSWORD', 'auth-service', '초기 Admin 비밀번호', 'Admin1234!', '-', '최초 로그인 후 변경 권장'],
    ['초기 비밀번호', 'DEFAULT_RESET_PASSWORD', 'auth-service', '비밀번호 초기화 값', 'eln0330', '-', '운영 시 변경 권장'],
    ['모니터링', 'OTEL_EXPORTER_OTLP_ENDPOINT', '모든 서비스', 'Jaeger 트레이스 수집', 'http://jaeger:4318', '-', 'Jaeger 미사용 시 제거 가능'],
]

for r, data in enumerate(envs, 5):
    for c, val in enumerate(data, 1):
        font = CODE_FONT if c in [2, 5] else (BOLD_FONT if c == 1 else None)
        fill = None
        if c == 6 and val == 'O (필수)':
            fill = LIGHT_RED
            font = RED_FONT
        data_cell(ws5, r, c, val, 'top', font, fill)
    ws5.row_dimensions[r].height = 45


# ================================================================
# 인쇄 설정
# ================================================================
for ws in wb.worksheets:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

wb.save('docs/troubleshooting/개발자_트러블슈팅.xlsx')
print('docs/troubleshooting/개발자_트러블슈팅.xlsx 생성 완료!')
