#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""운영자 트러블슈팅 가이드 Excel 생성"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 스타일 =====
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='맑은 고딕', size=16, bold=True, color='2F5496')
SUBTITLE_FONT = Font(name='맑은 고딕', size=12, bold=True, color='333333')
SECTION_FONT = Font(name='맑은 고딕', size=11, bold=True, color='2F5496')
NORMAL_FONT = Font(name='맑은 고딕', size=10)
BOLD_FONT = Font(name='맑은 고딕', size=10, bold=True)
SMALL_FONT = Font(name='맑은 고딕', size=9, color='555555')
CODE_FONT = Font(name='Consolas', size=9)
RED_FONT = Font(name='맑은 고딕', size=10, bold=True, color='C00000')
ORANGE_FONT = Font(name='맑은 고딕', size=10, bold=True, color='ED7D31')
GREEN_FONT = Font(name='맑은 고딕', size=10, bold=True, color='548235')

LIGHT_RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
LIGHT_ORANGE = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
LIGHT_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)


def header_row(ws, row, headers, widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[row].height = 28
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def data_cell(ws, row, col, value, align='left', font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or NORMAL_FONT
    cell.alignment = left if align == 'left' else (top_left if align == 'top' else center)
    cell.border = thin_border
    if fill:
        cell.fill = fill
    elif row % 2 == 0:
        cell.fill = LIGHT_GRAY
    return cell


def title(ws, text, row=1, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 40


def subtitle(ws, text, row, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.fill = LIGHT_BLUE
    cell.border = thin_border
    ws.row_dimensions[row].height = 28


# ================================================================
# 시트 1: 증상별 조치 가이드
# ================================================================
ws1 = wb.active
ws1.title = '증상별 조치 가이드'
ws1.sheet_properties.tabColor = 'C00000'

title(ws1, 'LabNote ELN 운영자 트러블슈팅 가이드 - 증상별 조치', cols=7)

ws1.merge_cells('A2:G2')
ws1['A2'].value = '사용자 또는 운영자가 경험하는 증상 기준으로 원인과 조치 방법을 안내합니다.'
ws1['A2'].font = SMALL_FONT
ws1['A2'].alignment = left

hdrs = ['No', '증상 (사용자가 보고하는 현상)', '영향 범위', '긴급도', '예상 원인', '조치 방법', '확인 방법']
widths = [5, 35, 12, 8, 30, 45, 30]
header_row(ws1, 4, hdrs, widths)

symptoms = [
    # ── 로그인/인증 ──
    [1, '로그인 버튼을 눌러도 반응이 없음\n(화면이 멈춤)', '전체 사용자', '긴급',
     'API Gateway(8000) 서비스 중단\n또는 네트워크 방화벽 차단',
     '1. docker ps 로 api-gateway 컨테이너 상태 확인\n2. 중단 시: docker compose up -d --build api-gateway\n3. 방화벽에서 8000번 포트 개방 확인',
     'http://서버IP:8000/health 접속하여\n{"status":"ok"} 응답 확인'],

    [2, '"이메일 또는 비밀번호가 올바르지 않습니다"\n오류가 반복됨', '특정 사용자', '보통',
     '1. 비밀번호 입력 오류\n2. 계정 비활성화 상태\n3. auth-service 장애',
     '1. 관리자 페이지(/admin/users)에서 해당 사용자 상태 확인\n2. 비활성(inactive) 시 → 활성화 처리\n3. 비밀번호 초기화 필요 시 관리자가 리셋',
     '해당 사용자 계정으로 다시 로그인 시도'],

    [3, '로그인은 되지만 페이지 이동 시\n"세션이 만료되었습니다" 반복 노출', '특정 사용자', '보통',
     '1. JWT 토큰 만료 (15분)\n2. Redis 세션 캐시 장애\n3. 서버 시간 동기화 오류',
     '1. 로그아웃 후 재로그인 안내\n2. docker compose logs redis 로 Redis 상태 확인\n3. Redis 재시작: docker compose restart redis\n4. 서버 시간 확인: date 명령어',
     '로그인 후 5분간 정상 사용 가능 여부 확인'],

    [4, '관리자 메뉴가 보이지 않음\n(권한 부족 안내)', '특정 사용자', '낮음',
     '사용자 역할이 admin이 아닌 상태',
     '1. 관리자 계정으로 접속\n2. 관리 > 사용자 관리에서 해당 사용자의 역할을 admin으로 변경\n※ 역할 변경 시 해당 사용자는 재로그인 필요',
     '해당 사용자가 재로그인 후 관리 메뉴 노출 확인'],

    # ── 연구노트 ──
    [5, '연구노트 목록이 빈 화면으로 나옴\n(로딩 후 아무것도 없음)', '전체 사용자', '긴급',
     '1. eln-service(8002) 중단\n2. PostgreSQL DB 연결 실패\n3. 데이터 미존재',
     '1. docker ps 로 eln-service 상태 확인\n2. docker compose logs eln-service 로 에러 확인\n3. DB 연결 오류 시: docker compose restart postgres\n4. eln-service 재시작: docker compose up -d --build eln-service',
     'http://서버IP:8000/api/notes 호출하여\n데이터 반환 여부 확인'],

    [6, '노트 저장 버튼을 눌러도\n"저장에 실패했습니다" 오류', '특정 사용자', '긴급',
     '1. 노트가 잠금(locked) 또는 서명(signed) 상태\n2. eln-service 장애\n3. DB 디스크 용량 부족',
     '1. 노트 상태 확인 (locked/signed 노트는 수정 불가 — 정상 동작)\n2. eln-service 로그 확인\n3. DB 디스크 확인: docker exec postgres df -h\n4. 디스크 부족 시 불필요 데이터 정리 또는 디스크 증설',
     '다른 draft 상태 노트에서 저장 테스트'],

    [7, '"이 상태 변경은 권한이 없습니다"\n오류 발생', '특정 사용자', '낮음',
     '역할별 상태 전환 제한 (정상 동작)\n- Researcher: 잠금(lock) 불가\n- 비Admin: 잠금 해제 불가',
     '※ 이것은 오류가 아니라 보안 정책입니다.\n- 잠금: Reviewer 또는 Admin 역할 필요\n- 잠금 해제: Admin 역할 + 비밀번호 확인 필요\n- 서명: Reviewer 또는 Admin만 가능',
     '해당 권한을 가진 사용자에게 요청 안내'],

    [8, '실시간 협업 편집이 안 됨\n("단독 편집 모드로 전환됩니다" 안내)', '특정 노트', '보통',
     '1. collab-service(8009) 중단\n2. WebSocket 연결 차단 (프록시/방화벽)\n3. Redis pub/sub 장애',
     '1. docker ps 로 collab-service 확인\n2. collab-service 재시작: docker compose up -d --build collab-service\n3. 방화벽에서 WebSocket(ws://) 프로토콜 허용 확인\n4. Redis 상태 확인',
     'http://서버IP:8009/health 접속하여\nrooms 수 확인'],

    # ── 전자서명 ──
    [9, '서명 버튼을 눌러도\n"서명에 실패했습니다" 오류', '특정 사용자', '긴급',
     '1. signature-audit-service(8003) 중단\n2. 비밀번호 불일치\n3. Redis Stream 연결 실패',
     '1. docker compose logs signature-audit-service 로 에러 확인\n2. 비밀번호 오류 시 올바른 비밀번호 입력 안내\n3. signature-audit-service 재시작\n4. Redis 상태 확인 및 재시작',
     '서명 재시도 후 노트 상태가 signed로 변경 확인'],

    [10, '서명은 완료됐는데\n노트 상태가 signed로 안 바뀜', '특정 노트', '긴급',
     'Redis Stream 이벤트 처리 실패\n(NOTE_SIGNED 이벤트 미소비)',
     '1. eln-service 로그에서 "event consumer" 관련 에러 확인\n2. Redis 상태 확인: docker compose logs redis\n3. eln-service 재시작 (재시작 시 미처리 이벤트 자동 복구)\n4. 그래도 안 되면 signature-audit-service도 재시작',
     'eln-service 로그에서 "NOTE_SIGNED processed" 메시지 확인\n노트 상태가 signed로 변경 확인'],

    # ── 파일 ──
    [11, '파일 업로드 시\n"업로드에 실패했습니다" 오류', '특정 사용자', '보통',
     '1. MinIO 스토리지 서비스 중단\n2. MinIO 디스크 용량 부족\n3. file-service(8008) 장애\n4. 차단 파일 형식 (.exe, .sh, .bat)',
     '1. docker ps 로 minio, file-service 상태 확인\n2. MinIO 콘솔(http://서버IP:9001) 접속하여 용량 확인\n3. file-service 재시작\n4. 차단 확장자 파일인 경우 사용자에게 안내',
     'http://서버IP:8008/health 접속\n소용량 텍스트 파일로 업로드 테스트'],

    [12, 'PDF/ZIP 내보내기가\n"변환 중..." 상태에서 멈춤', '특정 사용자', '보통',
     '1. BullMQ 작업 큐 정체\n2. signature-audit-service 과부하\n3. Redis 연결 문제',
     '1. docker compose logs signature-audit-service 에서 worker 로그 확인\n2. Redis 상태 확인\n3. signature-audit-service 재시작\n(재시작 시 미완료 작업은 자동 재처리)',
     '새로운 내보내기 작업 생성하여 정상 완료 확인'],

    # ── 검색 ──
    [13, '검색 결과가 하나도 안 나옴\n(검색 기능 자체가 동작 안 함)', '전체 사용자', '보통',
     '1. OpenSearch 서비스 중단\n2. search-service(8006) 장애\n3. 검색 인덱스 미생성',
     '1. docker ps 로 opensearch, search-service 확인\n2. OpenSearch 상태: curl http://서버IP:9200/_cluster/health\n3. search-service 재시작\n4. 인덱스 재생성 필요 시 search-service 재시작 (자동 생성)',
     'http://서버IP:9200/_cat/indices 로\nlab_search 인덱스 존재 여부 확인'],

    [14, '최근 작성한 노트가\n검색 결과에 안 나옴', '특정 사용자', '낮음',
     '1. 검색 인덱스 동기화 지연\n2. Redis 캐시에 이전 결과 남아있음',
     '1. 일반적으로 1~2분 내 자동 동기화됨 (대기)\n2. 계속 안 나오면 search-service 재시작\n(재시작 시 캐시 초기화 + 인덱스 재확인)',
     '2~3분 후 다시 검색하여 결과 확인'],

    # ── 인벤토리 ──
    [15, '인벤토리 항목 등록 시\n"이미 등록된 바코드입니다" 오류', '특정 사용자', '낮음',
     '동일 바코드가 이미 시스템에 등록됨 (정상 동작)',
     '※ 바코드는 조직 내 유일해야 합니다.\n1. 기존 등록된 항목 검색하여 확인\n2. 다른 바코드 사용 또는 기존 항목 수정',
     '해당 바코드로 검색하여 기존 항목 확인'],

    [16, '재고 출고 시\n"재고가 부족합니다" 오류', '특정 사용자', '낮음',
     '요청 수량 > 현재 재고 수량 (정상 동작)',
     '※ 시스템이 음수 재고를 방지하는 정상 동작입니다.\n1. 현재 재고 수량 확인\n2. 재고 수량 이하로 출고 요청\n3. 필요 시 입고(IN) 처리 후 출고',
     '해당 항목의 현재 재고 수량 확인'],

    # ── 예약 ──
    [17, '예약 시\n"해당 시간대에 이미 예약이 있습니다" 오류', '특정 사용자', '낮음',
     '동일 자원에 시간이 겹치는 예약 존재 (정상 동작)',
     '※ 이중 예약 방지를 위한 정상 동작입니다.\n1. 캘린더에서 해당 자원의 예약 현황 확인\n2. 비어있는 시간대로 변경하여 재예약',
     '스케줄러 캘린더에서 해당 시간대 확인'],

    # ── 시스템 전체 ──
    [18, '모든 페이지에서\n"서버와 통신할 수 없습니다" 오류', '전체 사용자', '긴급',
     '1. API Gateway(8000) 완전 중단\n2. 서버 자체 다운\n3. 네트워크 단절',
     '1. 서버 접속 가능 여부 확인 (SSH/원격접속)\n2. docker ps 로 전체 컨테이너 상태 확인\n3. 전체 재시작: docker compose up -d\n4. 네트워크 확인: ping, telnet 서버IP 8000',
     'docker ps 로 모든 컨테이너 Up 상태 확인\nhttp://서버IP:8000/health 접속 확인'],

    [19, '특정 시간대에 시스템이 느려짐\n(응답 지연)', '전체 사용자', '보통',
     '1. DB 커넥션 풀 고갈\n2. Redis 메모리 한도(128MB) 초과\n3. 서버 리소스(CPU/메모리) 부족',
     '1. 서버 리소스 확인: docker stats\n2. PostgreSQL 커넥션 확인: docker exec postgres psql -U labnote -c "SELECT count(*) FROM pg_stat_activity"\n3. Redis 메모리: docker exec redis redis-cli info memory\n4. 필요 시 서비스 순차 재시작',
     'docker stats 로 CPU/메모리 사용률 확인\n응답 시간 정상화 여부 확인'],

    [20, '대시보드 통계 일부가\n빈칸으로 표시됨', '전체 사용자', '낮음',
     '특정 서비스 장애 시 해당 통계만 null 반환\n(시스템이 장애를 격리하는 정상 동작)',
     '※ 대시보드는 일부 서비스 장애 시에도 나머지를 표시하도록 설계됨\n1. 빈칸 항목에 해당하는 서비스 로그 확인\n2. 해당 서비스 재시작\n3. 재시작 후 대시보드 새로고침',
     '빈칸이었던 통계가 정상 표시되는지 확인'],
]

for r, data in enumerate(symptoms, 5):
    data_cell(ws1, r, 1, data[0], 'center')
    data_cell(ws1, r, 2, data[1], 'top')
    data_cell(ws1, r, 3, data[2], 'center')
    # 긴급도 색상
    urgency = data[3]
    if urgency == '긴급':
        data_cell(ws1, r, 4, urgency, 'center', RED_FONT, LIGHT_RED)
    elif urgency == '보통':
        data_cell(ws1, r, 4, urgency, 'center', ORANGE_FONT, LIGHT_ORANGE)
    else:
        data_cell(ws1, r, 5, urgency, 'center', GREEN_FONT, LIGHT_GREEN)
        data_cell(ws1, r, 4, urgency, 'center', GREEN_FONT, LIGHT_GREEN)
    data_cell(ws1, r, 5, data[4], 'top')
    data_cell(ws1, r, 6, data[5], 'top')
    data_cell(ws1, r, 7, data[6], 'top')
    ws1.row_dimensions[r].height = 80


# ================================================================
# 시트 2: 서비스 상태 점검
# ================================================================
ws2 = wb.create_sheet('서비스 상태 점검')
ws2.sheet_properties.tabColor = '4472C4'

title(ws2, '서비스 상태 점검 가이드', cols=7)

ws2.merge_cells('A2:G2')
ws2['A2'].value = '각 서비스의 정상 동작 여부를 확인하는 방법입니다. 장애 발생 시 아래 순서대로 점검하세요.'
ws2['A2'].font = SMALL_FONT
ws2['A2'].alignment = left

# 2-1: 전체 서비스 목록
subtitle(ws2, '1. 서비스 구성 및 점검 URL', 4, 7)

hdrs = ['No', '서비스명', '설명 (역할)', '포트', '상태 확인 URL', '상태 확인 명령어', '정상 응답']
widths = [5, 18, 25, 8, 32, 35, 20]
header_row(ws2, 5, hdrs, widths)

services = [
    [1, 'API Gateway', '모든 요청의 진입점\nJWT 인증, 프록시, Rate Limit', 8000,
     'http://서버IP:8000/health', 'curl http://localhost:8000/health', '{"status":"ok"}'],
    [2, 'auth-service', '로그인, 회원가입\n사용자/팀/역할 관리', 8001,
     'http://서버IP:8001/health', 'docker exec auth-service curl -s http://localhost:8001/health', '{"status":"ok"}'],
    [3, 'eln-service', '연구노트 작성/편집\n상태관리, 템플릿', 8002,
     'http://서버IP:8002/health', 'docker exec eln-service curl -s http://localhost:8002/health', '{"status":"ok"}'],
    [4, 'signature-audit', '전자서명, 감사로그\nPDF/ZIP 내보내기, 알림', 8003,
     'http://서버IP:8003/health', 'docker exec signature-audit-service curl -s http://localhost:8003/health', '{"status":"ok"}'],
    [5, 'inventory-service', '시약/장비 관리\n바코드, 수량 입출고', 8004,
     'http://서버IP:8004/health', 'docker exec inventory-service curl -s http://localhost:8004/health', '{"status":"ok"}'],
    [6, 'scheduler-service', '장비/회의실 예약\n승인/반려 관리', 8005,
     'http://서버IP:8005/health', 'docker exec scheduler-service curl -s http://localhost:8005/health', '{"status":"ok"}'],
    [7, 'search-service', '통합 검색\n자동완성, 이력 관리', 8006,
     'http://서버IP:8006/health', 'docker exec search-service curl -s http://localhost:8006/health', '{"status":"ok"}'],
    [8, 'file-service', '파일 업/다운로드\nMinIO 연동', 8008,
     'http://서버IP:8008/health', 'docker exec file-service curl -s http://localhost:8008/health', '{"status":"ok"}'],
    [9, 'collab-service', '실시간 협업 편집\nWebSocket, 커서 공유', 8009,
     'http://서버IP:8009/health', 'docker exec collab-service curl -s http://localhost:8009/health', '{"status":"ok","rooms":N}'],
    [10, 'PostgreSQL', '메인 데이터베이스', 5432,
     '-', 'docker exec postgres pg_isready', 'accepting connections'],
    [11, 'Redis', '캐시, 세션, 이벤트 스트림', 6379,
     '-', 'docker exec redis redis-cli ping', 'PONG'],
    [12, 'MinIO', '파일 저장소 (S3 호환)', '9000\n(콘솔:9001)',
     'http://서버IP:9001 (콘솔)', 'curl http://localhost:9000/minio/health/live', 'HTTP 200'],
    [13, 'OpenSearch', '검색 엔진\n한국어 형태소 분석', 9200,
     'http://서버IP:9200', 'curl http://localhost:9200/_cluster/health', '"status":"green/yellow"'],
    [14, 'Jaeger', '요청 추적 (트레이싱)\n성능 분석', 16686,
     'http://서버IP:16686', '브라우저에서 접속', 'Jaeger UI 화면'],
    [15, 'Dozzle', '실시간 로그 뷰어', 9999,
     'http://서버IP:9999', '브라우저에서 접속', 'Dozzle 로그 화면'],
]

for r, data in enumerate(services, 6):
    for c, val in enumerate(data, 1):
        align = 'top' if c in [3, 5, 6] else 'center'
        font = CODE_FONT if c in [5, 6, 7] else None
        data_cell(ws2, r, c, val, align, font)
    ws2.row_dimensions[r].height = 40

# 2-2: 빠른 전체 점검 명령어
r = 6 + len(services) + 1
subtitle(ws2, '2. 빠른 전체 점검 명령어', r, 7)
r += 1
commands = [
    ['전체 컨테이너 상태 확인', 'docker ps --format "table {{.Names}}\\t{{.Status}}\\t{{.Ports}}"', '모든 컨테이너가 Up 상태'],
    ['리소스 사용량 확인', 'docker stats --no-stream', 'CPU/메모리 사용률 확인'],
    ['전체 서비스 로그 (최근 50줄)', 'docker compose logs --tail=50', '에러(ERROR) 메시지 없음'],
    ['특정 서비스 로그 확인', 'docker compose logs -f <서비스명>', '실시간 로그 모니터링'],
    ['디스크 사용량 확인', 'df -h && docker system df', '용량 80% 미만'],
    ['DB 커넥션 수 확인', 'docker exec postgres psql -U labnote -c\n"SELECT count(*) FROM pg_stat_activity"', '30 미만 (max: 30)'],
    ['Redis 메모리 확인', 'docker exec redis redis-cli info memory | grep used_memory_human', '128MB 미만'],
]

hdrs2 = ['목적', '명령어', '정상 기준', '', '', '', '']
header_row(ws2, r, hdrs2)
ws2.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

for i, cmd in enumerate(commands, r + 1):
    data_cell(ws2, i, 1, cmd[0], 'left', BOLD_FONT)
    ws2.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
    data_cell(ws2, i, 2, cmd[1], 'left', CODE_FONT)
    ws2.merge_cells(start_row=i, start_column=6, end_row=i, end_column=7)
    data_cell(ws2, i, 6, cmd[2], 'left')
    ws2.row_dimensions[i].height = 32


# ================================================================
# 시트 3: 서비스 재시작 절차
# ================================================================
ws3 = wb.create_sheet('서비스 재시작 절차')
ws3.sheet_properties.tabColor = 'ED7D31'

title(ws3, '서비스 재시작 절차', cols=6)

ws3.merge_cells('A2:F2')
ws3['A2'].value = '서비스 재시작 시 의존성 순서를 반드시 지켜야 합니다. 잘못된 순서로 재시작하면 연쇄 장애가 발생할 수 있습니다.'
ws3['A2'].font = Font(name='맑은 고딕', size=10, color='C00000', bold=True)
ws3['A2'].alignment = left

# 3-1: 의존성 순서
subtitle(ws3, '1. 서비스 기동 순서 (반드시 이 순서대로)', 4, 6)

hdrs = ['순서', '서비스', '의존 대상', '재시작 명령어', '대기 시간', '확인']
widths = [6, 18, 22, 40, 10, 25]
header_row(ws3, 5, hdrs, widths)

restart_order = [
    ['1단계', 'PostgreSQL', '없음 (최우선)', 'docker compose up -d postgres', '30초', 'pg_isready 응답 확인'],
    ['1단계', 'Redis', '없음 (최우선)', 'docker compose up -d redis', '10초', 'redis-cli ping → PONG'],
    ['1단계', 'MinIO', '없음 (최우선)', 'docker compose up -d minio', '15초', '콘솔(9001) 접속 확인'],
    ['1단계', 'OpenSearch', '없음 (최우선)', 'docker compose up -d opensearch', '60초', '9200 포트 응답 확인'],
    ['2단계', 'auth-service', 'PostgreSQL, Redis', 'docker compose up -d --build auth-service', '15초', '/health 응답 확인'],
    ['2단계', 'eln-service', 'PostgreSQL, Redis', 'docker compose up -d --build eln-service', '15초', '/health 응답 확인'],
    ['2단계', 'search-service', 'PostgreSQL, OpenSearch, Redis', 'docker compose up -d --build search-service', '15초', '/health 응답 확인'],
    ['2단계', 'file-service', 'PostgreSQL, MinIO', 'docker compose up -d --build file-service', '15초', '/health 응답 확인'],
    ['2단계', 'inventory-service', 'PostgreSQL, Redis', 'docker compose up -d --build inventory-service', '15초', '/health 응답 확인'],
    ['2단계', 'scheduler-service', 'PostgreSQL', 'docker compose up -d --build scheduler-service', '15초', '/health 응답 확인'],
    ['2단계', 'signature-audit', 'PostgreSQL, Redis', 'docker compose up -d --build signature-audit-service', '15초', '/health 응답 확인'],
    ['2단계', 'collab-service', 'Redis', 'docker compose up -d --build collab-service', '10초', '/health 응답 확인'],
    ['3단계', 'API Gateway', '모든 백엔드 서비스', 'docker compose up -d --build api-gateway', '10초', '/health 응답 확인'],
]

for r, data in enumerate(restart_order, 6):
    stage = data[0]
    fill = None
    if '1단계' in stage:
        fill = LIGHT_BLUE
    elif '3단계' in stage:
        fill = LIGHT_ORANGE
    for c, val in enumerate(data, 1):
        font = CODE_FONT if c == 4 else None
        data_cell(ws3, r, c, val, 'center' if c in [1, 5] else 'left', font, fill)
    ws3.row_dimensions[r].height = 28

# 3-2: 전체 재시작
r2 = 6 + len(restart_order) + 1
subtitle(ws3, '2. 전체 시스템 재시작 (한 번에)', r2, 6)
r2 += 1
full_restart = [
    ['전체 중지', 'docker compose down', '모든 컨테이너 중지 (데이터는 유지됨)'],
    ['전체 시작', 'docker compose up -d', '의존성 순서에 따라 자동 기동'],
    ['전체 빌드+시작', 'docker compose up -d --build', '코드 변경 반영이 필요한 경우'],
    ['상태 확인', 'docker ps', '모든 컨테이너 Up + healthy 상태 확인'],
]

hdrs = ['작업', '명령어', '설명', '', '', '']
header_row(ws3, r2, hdrs)
for i, data in enumerate(full_restart, r2 + 1):
    data_cell(ws3, i, 1, data[0], 'left', BOLD_FONT)
    ws3.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    data_cell(ws3, i, 2, data[1], 'left', CODE_FONT)
    ws3.merge_cells(start_row=i, start_column=5, end_row=i, end_column=6)
    data_cell(ws3, i, 5, data[2], 'left')
    ws3.row_dimensions[i].height = 28

# 3-3: 주의사항
r3 = r2 + len(full_restart) + 2
subtitle(ws3, '3. 재시작 시 주의사항', r3, 6)
r3 += 1
cautions = [
    ['docker compose down -v 금지', '볼륨(-v)까지 삭제하면 DB, 파일 등 모든 데이터가 영구 삭제됩니다.', '긴급'],
    ['서비스 순서 준수', 'DB/Redis가 준비되기 전에 백엔드 서비스를 시작하면 연결 실패합니다.', '긴급'],
    ['포트 충돌 확인', '다른 프로그램이 같은 포트를 사용 중이면 서비스가 시작되지 않습니다.', '보통'],
    ['디스크 용량 확인', '디스크 부족 시 PostgreSQL, MinIO 등이 정상 동작하지 않습니다.', '보통'],
    ['로그 확인 습관', '재시작 후 30초간 docker compose logs -f 로 에러 여부를 확인하세요.', '권장'],
]

hdrs = ['주의사항', '설명', '중요도', '', '', '']
header_row(ws3, r3, hdrs)
for i, data in enumerate(cautions, r3 + 1):
    data_cell(ws3, i, 1, data[0], 'left', BOLD_FONT)
    ws3.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
    data_cell(ws3, i, 2, data[1], 'left')
    urg = data[2]
    fill = LIGHT_RED if urg == '긴급' else (LIGHT_ORANGE if urg == '보통' else LIGHT_GREEN)
    font = RED_FONT if urg == '긴급' else (ORANGE_FONT if urg == '보통' else GREEN_FONT)
    data_cell(ws3, i, 6, urg, 'center', font, fill)
    ws3.row_dimensions[i].height = 28


# ================================================================
# 시트 4: 데이터 백업/복구
# ================================================================
ws4 = wb.create_sheet('데이터 백업 및 복구')
ws4.sheet_properties.tabColor = '548235'

title(ws4, '데이터 백업 및 복구 가이드', cols=6)

ws4.merge_cells('A2:F2')
ws4['A2'].value = '정기 백업과 장애 시 복구 절차를 안내합니다. 백업은 최소 주 1회 이상 수행하는 것을 권장합니다.'
ws4['A2'].font = SMALL_FONT
ws4['A2'].alignment = left

# 4-1: 데이터 저장 위치
subtitle(ws4, '1. 데이터 저장 위치 (Docker Volume)', 4, 6)

hdrs = ['구분', '볼륨명', '저장 내용', '백업 필수', '비고', '']
widths = [12, 18, 25, 10, 25, 10]
header_row(ws4, 5, hdrs, widths)

volumes = [
    ['데이터베이스', 'postgres_data', '모든 서비스의 업무 데이터\n(사용자, 노트, 서명, 인벤토리 등)', '필수',
     '가장 중요 — 이 데이터 없이는 복구 불가'],
    ['파일 저장소', 'minio_data', '업로드된 파일, 내보낸 PDF/ZIP', '필수',
     '연구 데이터 원본 파일 포함'],
    ['캐시/이벤트', 'redis_data', '세션, 캐시, 이벤트 스트림', '선택',
     '손실 시 재로그인 필요\n캐시는 자동 재생성'],
    ['검색 인덱스', 'opensearch_data', '검색용 인덱스 데이터', '선택',
     '손실 시 search-service\n재시작으로 자동 재생성'],
]

for r, data in enumerate(volumes, 6):
    for c, val in enumerate(data, 1):
        font = RED_FONT if c == 4 and val == '필수' else None
        fill2 = LIGHT_RED if c == 4 and val == '필수' else None
        data_cell(ws4, r, c, val, 'top' if c in [3, 5] else 'center', font, fill2)
    ws4.row_dimensions[r].height = 45

# 4-2: 백업 절차
r = 6 + len(volumes) + 1
subtitle(ws4, '2. 백업 절차', r, 6)
r += 1

backups = [
    ['PostgreSQL\n(가장 중요)', '전체 백업 (dump)',
     'docker exec postgres pg_dumpall -U labnote > backup_$(date +%Y%m%d).sql',
     '주 1회 이상\n(일 1회 권장)',
     '백업 파일을 외부 저장소(NAS 등)에 보관\nSQL 파일 크기로 정상 여부 판단'],
    ['PostgreSQL', '특정 DB만 백업',
     'docker exec postgres pg_dump -U labnote labnote > labnote_$(date +%Y%m%d).sql',
     '필요 시',
     '특정 데이터베이스만 백업할 때'],
    ['MinIO (파일)', '볼륨 복사',
     'docker run --rm -v minio_data:/data -v $(pwd):/backup alpine tar czf /backup/minio_$(date +%Y%m%d).tar.gz /data',
     '주 1회',
     '파일 용량에 따라 시간 소요\n업무 외 시간 권장'],
    ['Redis', 'RDB 스냅샷 복사',
     'docker exec redis redis-cli BGSAVE && docker cp redis:/data/dump.rdb ./redis_backup_$(date +%Y%m%d).rdb',
     '선택',
     '손실되어도 재로그인으로 복구 가능'],
    ['Docker Volume\n(전체)', '볼륨 일괄 백업',
     '각 볼륨별 tar 명령 실행\n(위 방법들을 순차 실행)',
     '월 1회',
     '전체 시스템 이전 시 사용'],
]

hdrs = ['대상', '방법', '명령어', '주기', '비고', '']
header_row(ws4, r, hdrs)
for i, data in enumerate(backups, r + 1):
    for c, val in enumerate(data, 1):
        font = CODE_FONT if c == 3 else (BOLD_FONT if c == 1 else None)
        data_cell(ws4, i, c, val, 'top', font)
    ws4.row_dimensions[i].height = 55

# 4-3: 복구 절차
r2 = r + len(backups) + 2
subtitle(ws4, '3. 복구 절차', r2, 6)
r2 += 1

restores = [
    ['PostgreSQL 복구', '1. 서비스 중지: docker compose stop (postgres 제외)\n2. 기존 DB 삭제: docker exec postgres psql -U labnote -c "DROP DATABASE labnote"\n3. DB 재생성: docker exec postgres psql -U labnote -c "CREATE DATABASE labnote"\n4. 백업 복원: cat backup.sql | docker exec -i postgres psql -U labnote\n5. 서비스 재시작: docker compose up -d',
     '모든 서비스 중지 후 복구\n복구 후 전체 서비스 재시작 필수'],
    ['MinIO 복구', '1. file-service 중지: docker compose stop file-service\n2. 볼륨 복원: docker run --rm -v minio_data:/data -v $(pwd):/backup alpine tar xzf /backup/minio_backup.tar.gz -C /\n3. 서비스 재시작: docker compose up -d file-service',
     '파일 용량에 따라 복원 시간 상이'],
    ['Redis 복구', '1. redis 중지: docker compose stop redis\n2. RDB 파일 복사: docker cp redis_backup.rdb redis:/data/dump.rdb\n3. redis 재시작: docker compose up -d redis',
     '선택 사항 — 손실 시 재로그인만 필요'],
    ['전체 시스템 이전', '1. 새 서버에 Docker, Docker Compose 설치\n2. 프로젝트 코드 복사\n3. .env 파일 설정\n4. 백업 파일들 복사\n5. 위 복구 절차 순서대로 실행\n6. docker compose up -d',
     '새 서버 이전 시 사용\n.env 파일 반드시 포함'],
]

hdrs = ['대상', '절차', '비고', '', '', '']
header_row(ws4, r2, hdrs)
for i, data in enumerate(restores, r2 + 1):
    data_cell(ws4, i, 1, data[0], 'top', BOLD_FONT)
    ws4.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    data_cell(ws4, i, 2, data[1], 'top', CODE_FONT)
    ws4.merge_cells(start_row=i, start_column=5, end_row=i, end_column=6)
    data_cell(ws4, i, 5, data[2], 'top')
    ws4.row_dimensions[i].height = 100


# ================================================================
# 시트 5: 자주 묻는 질문 (FAQ)
# ================================================================
ws5 = wb.create_sheet('자주 묻는 질문(FAQ)')
ws5.sheet_properties.tabColor = '7030A0'

title(ws5, '자주 묻는 질문 (FAQ)', cols=5)

hdrs = ['No', '질문', '답변', '관련 기능', '참고']
widths = [5, 35, 55, 14, 20]
header_row(ws5, 3, hdrs, widths)

faqs = [
    [1, '시스템을 처음 설치한 후\n기본 관리자 계정은?',
     '이메일: admin@labnote.com\n비밀번호: Admin1234!\n※ 최초 로그인 후 반드시 비밀번호를 변경하세요.',
     '인증', '초기 시드 데이터로\n자동 생성됨'],
    [2, '서명(signed) 된 노트를\n수정하거나 삭제할 수 있나요?',
     '아니요. 서명된 노트는 위변조 방지를 위해\n어떤 역할도 수정/삭제할 수 없습니다.\n이것은 전자연구노트의 법적 요건입니다.',
     '연구노트', '법적 요건에 의한\n의도된 제한'],
    [3, '잠금(locked) 된 노트를\n다시 편집하려면?',
     'Admin 역할을 가진 사용자가\n노트 상세 → "잠금 해제" 버튼 클릭 → 비밀번호 입력\n→ 잠금 해제 후 draft 상태로 복귀합니다.',
     '연구노트', 'Admin 전용 기능'],
    [4, '사용자의 비밀번호를\n초기화하려면?',
     'Admin이 관리 > 사용자 관리에서\n해당 사용자 선택 → 비밀번호 리셋\n초기화 비밀번호: eln0330',
     '인증', '사용자에게 변경 안내 필요'],
    [5, '특정 서비스만 재시작하면\n다른 서비스에 영향이 있나요?',
     '대부분의 서비스는 독립적으로 재시작 가능합니다.\n단, PostgreSQL이나 Redis를 재시작하면\n의존하는 모든 서비스에 일시적 영향이 있습니다.\n(자동 재연결되므로 보통 10~30초 내 복구)',
     '인프라', '서비스 재시작 절차\n시트 참고'],
    [6, '디스크 용량이 부족하면\n어떻게 해야 하나요?',
     '1. docker system prune -a 로 미사용 이미지/캐시 정리\n2. 오래된 Export 파일 정리 (MinIO 콘솔에서)\n3. PostgreSQL VACUUM 실행\n4. 근본적으로는 디스크 증설 필요',
     '인프라', 'prune은 미사용\n이미지만 삭제'],
    [7, '검색이 느리거나\n결과가 부정확하면?',
     '1. OpenSearch 상태 확인: curl http://localhost:9200/_cluster/health\n2. search-service 재시작으로 인덱스 갱신\n3. 한국어 검색은 형태소 분석기(nori) 기반으로\n   완벽한 부분 일치는 지원하지 않을 수 있음',
     '검색', 'OpenSearch 2.x\nnori 분석기 사용'],
    [8, '동시에 여러 명이\n같은 노트를 편집할 수 있나요?',
     '네, 실시간 협업 편집을 지원합니다.\nWebSocket 기반으로 커서 위치와 내용이 실시간 동기화됩니다.\n단, collab-service와 Redis가 정상이어야 합니다.',
     '협업', 'collab-service\n(포트 8009)'],
    [9, '시스템 로그를\n실시간으로 보려면?',
     '1. Dozzle 로그 뷰어: http://서버IP:9999\n   (브라우저에서 바로 확인 가능)\n2. 터미널: docker compose logs -f <서비스명>\n3. Jaeger 트레이싱: http://서버IP:16686',
     '모니터링', '비밀번호 없이 접속 가능\n내부망에서만 사용 권장'],
    [10, '서버 재부팅 후\n시스템이 자동으로 시작되나요?',
     'Docker Compose 서비스에 restart: unless-stopped 설정이 되어 있으면\n서버 재부팅 시 자동 시작됩니다.\n확인: docker compose config | grep restart\n미설정 시: docker compose up -d 로 수동 시작',
     '인프라', 'docker compose.yml\n설정 확인'],
    [11, '외부에서 접속하려면\n어떤 포트를 열어야 하나요?',
     '필수: 8000 (API Gateway) + 프론트엔드 포트(5173 또는 배포 포트)\n선택: 9001 (MinIO 콘솔), 16686 (Jaeger), 9999 (Dozzle)\n※ DB(5432), Redis(6379) 등은 절대 외부 개방 금지!',
     '네트워크', '보안상 최소한의\n포트만 개방'],
    [12, '조직(Organization)이란\n무엇인가요?',
     '데이터 격리 단위입니다. 각 조직의 사용자는\n자기 조직의 데이터만 볼 수 있습니다.\n다른 조직의 노트, 인벤토리, 예약 등은 접근 불가합니다.\n(멀티테넌시 보안)',
     '인증/보안', '조직 간 데이터\n완전 격리'],
]

for r, data in enumerate(faqs, 4):
    for c, val in enumerate(data, 1):
        align = 'top' if c in [2, 3, 5] else 'center'
        data_cell(ws5, r, c, val, align)
    ws5.row_dimensions[r].height = 75


# ================================================================
# 인쇄 설정
# ================================================================
for ws in wb.worksheets:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False

wb.save('docs/troubleshooting/운영자_트러블슈팅.xlsx')
print('docs/troubleshooting/운영자_트러블슈팅.xlsx 생성 완료!')
